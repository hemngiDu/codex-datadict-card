#!/usr/bin/env python3
"""Generate graph.html - all data inline, works with file://"""
import openpyxl, os, sys, json

DG = [("配置与方案",["库存计划方案","安全库存统计方案","日均消耗统计方案","调度方案"]),("因子与水位",["库存水位信息","库存水位因子","因子数据维护","因子取数方案","因子取数规则","补货日期"]),("算法与计算",["算法方案配置","算法注册配置","库存水位更新方案","智能计算设置","因子分批计算参数","快速搭建模型","资源注册模型"]),("执行与结果",["库存计划建议","供需匹配明细","安全库存记录","匹配映射配置","日均消耗记录","客户服务水平"]),("日志审计",["消耗量计算日志","安全库存计算日志","库存计划运算日志"]),("辅助测试",["组织测试"])]
CR = [("t_invp_scheme","t_invp_model_register","需求/供应来源模型"),("t_invp_scheme","t_invp_algoconfig","算法方案"),("t_invp_scheme","t_invp_calrulecfg","水位更新方案"),("t_invp_scheme","t_invp_invlevel","库存水位信息"),("t_invp_scheme","t_invp_matchconfig","匹配映射配置"),("t_invp_planadvice","t_invp_scheme","所属计划方案"),("t_invp_invlevel","t_invp_levelfactor","水位因子"),("t_invp_matchdetail","t_invp_scheme","计划方案"),("t_invp_smartcalccfg","t_invp_invlevel","库存水位"),("t_invp_smartcalccfg","t_invp_calrulecfg","因子计算规则"),("t_invp_smartcalccfg","t_invp_queryschema","因子取数方案"),("t_invp_invleveldata","t_invp_levelfactor","水位因子"),("t_invp_ssrecord","t_invp_ssdayscheme","安全库存方案"),("t_invp_dac_record","t_invp_dailycomsumption","日均消耗方案"),("t_invp_safestock_callog","t_invp_ssdayscheme","安全库存方案")]
def gd(bo):
    for d,objs in DG:
        if bo in objs: return d
    return "其他"

def main():
    if len(sys.argv)<2:
        print("Usage: python gen.py <Excel>"); sys.exit(1)
    path=sys.argv[1]
    if os.path.isdir(path):
        fs=[f for f in os.listdir(path) if f.endswith(".xlsx") and not f.startswith("~$")]
        if not fs: print("No xlsx"); sys.exit(1)
        path=os.path.join(path,fs[0])
    outdir=os.path.dirname(os.path.abspath(path))
    wb=openpyxl.load_workbook(path)
    ws=wb["表目录"]
    
    rows=[];cn=cb=None
    for i in range(7,ws.max_row+1):
        c4=ws.cell(i,4).value;c5=ws.cell(i,5).value;c6=ws.cell(i,6).value;c7=ws.cell(i,7).value;c8=ws.cell(i,8).value
        if c4:
            try: cn=int(str(c4).strip())
            except: pass
        if c5: cb=str(c5).strip()
        if c6 and c7: rows.append({"bo":cb,"desc":str(c6).strip(),"tbl":str(c7).strip(),"rel":str(c8).strip() if c8 else ""})
    
    nids=set();nodes=[];edges=[]
    for row in rows:
        if row["tbl"]in nids: continue
        nids.add(row["tbl"])
        nodes.append({"id":row["tbl"],"label":row["tbl"].replace("t_invp_","").replace("tk_",""),"bo":row["bo"],"desc":row["desc"],"domain":gd(row["bo"]),"type":"main"if not row["rel"]else"child"})
    for row in rows:
        if row["rel"]and row["rel"]!=row["tbl"]and row["tbl"]in nids and row["rel"]in nids:
            edges.append({"source":row["rel"],"target":row["tbl"],"label":row["desc"]})
    for s,t,l in CR:
        if s in nids and t in nids: edges.append({"source":s,"target":t,"label":l,"type":"cross"})
    
    cols={}
    for sn in wb.sheetnames:
        if sn=="表目录": continue
        s=wb[sn];ct=None;c=[];h=False
        for i in range(1,s.max_row+1):
            c2=s.cell(i,2).value;c5=s.cell(i,5).value;c8=s.cell(i,8).value
            if c2 and str(c2).strip()=="数据表名：" and c5:
                if ct and c:
                    if ct not in cols or len(c)>len(cols.get(ct,[])): cols[ct]=c
                ct=str(c5).strip();c=[];h=False
            elif c2 and str(c2).strip()=="编号":
                c3=s.cell(i,3).value;c4=s.cell(i,4).value
                if c3 and c4 and str(c3).strip()=="列名" and "列标题" in str(c4): h=True
            elif h and c2:
                try:
                    float(str(c2).strip())
                    cn2=str(s.cell(i,3).value).strip() if s.cell(i,3).value else ""
                    ct2=str(s.cell(i,4).value).strip() if s.cell(i,4).value else cn2
                    tp=str(s.cell(i,5).value).strip() if s.cell(i,5).value else ""
                    rr=str(s.cell(i,11).value).strip() if s.cell(i,11).value else ""
                    if cn2: c.append({"n":cn2,"t":ct2,"tp":tp,"r":rr})
                except:
                    if c8 and str(c8).strip()=="返回主目录": h=False
        if ct and c:
            if ct not in cols or len(c)>len(cols.get(ct,[])): cols[ct]=c
    
    j_idx=json.dumps([{"id":n["id"],"label":n["label"],"bo":n["bo"],"domain":n["domain"],"desc":n["desc"]} for n in nodes],ensure_ascii=False)
    j_nm=json.dumps({n["id"]:n for n in nodes},ensure_ascii=False)
    j_eg=json.dumps(edges,ensure_ascii=False)
    j_cl=json.dumps(cols,ensure_ascii=False)
    j_cr=json.dumps({"配置与方案":"#3b82f6","因子与水位":"#10b981","算法与计算":"#f59e0b","执行与结果":"#ef4444","日志审计":"#6b7280","辅助测试":"#8b5cf6","其他":"#94a3b8"},ensure_ascii=False)
    
    # Build HTML piece by piece (avoid f-string escaping issues)
    parts=[]
    parts.append('''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>数据字典 · 关联模型</title>
<script src="https://d3js.org/d3.v7.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0f172a;color:#e2e8f0;overflow:hidden}
#hd{position:fixed;top:0;left:0;right:0;z-index:10;padding:12px 24px;background:rgba(15,23,42,0.9);backdrop-filter:blur(12px);border-bottom:1px solid rgba(255,255,255,0.06);display:flex;align-items:center;gap:12px}
#hd h1{font-size:18px;font-weight:600}#hd span{font-size:13px;color:#94a3b8}
#sr{position:fixed;top:56px;left:50%;transform:translateX(-50%);z-index:15;width:420px;padding:10px 0}
#sr input{width:100%;padding:9px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.08);background:rgba(30,41,59,0.95);color:#e2e8f0;font-size:13px;outline:none;transition:border-color 0.2s;box-sizing:border-box}
#sr input:focus{border-color:rgba(255,255,255,0.25)}#sr input::placeholder{color:#475569}
#sg{position:absolute;top:48px;left:0;right:0;background:#1e293b;border:1px solid rgba(255,255,255,0.08);border-radius:8px;max-height:320px;overflow-y:auto;display:none}
.si{padding:8px 14px;cursor:pointer;font-size:13px;border-bottom:1px solid rgba(255,255,255,0.03);display:flex;align-items:center;gap:8px;transition:background 0.15s}
.si:hover{background:rgba(255,255,255,0.06)}.si .sl{font-family:monospace;color:#e2e8f0}.si .sb{color:#64748b;font-size:11px}
.si .sd{font-size:10px;padding:1px 6px;border-radius:3px;margin-left:auto}
#lg{position:fixed;bottom:24px;left:24px;z-index:10;display:flex;gap:8px;flex-wrap:wrap;background:rgba(30,41,59,0.9);backdrop-filter:blur(12px);padding:10px 14px;border-radius:10px;border:1px solid rgba(255,255,255,0.06)}
.li{display:flex;align-items:center;gap:6px;font-size:12px;color:#cbd5e1}.ld{width:10px;height:10px;border-radius:50%}
#tp{z-index:20;background:rgba(30,41,59,0.96);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:14px 18px;font-size:13px;line-height:1.6;max-width:400px;pointer-events:none;opacity:0;transition:opacity 0.15s;position:fixed;z-index:20}
#tp.show{opacity:1}#tp .tt{font-size:15px;font-weight:600;color:#f1f5f9;word-break:break-all}#tp .tb{color:#94a3b8;font-size:12px}
svg{position:fixed;top:0;left:0;width:100%;height:100%}
.el{stroke:rgba(255,255,255,0.12);stroke-width:1.5;fill:none;transition:stroke 0.2s}
.ec{stroke:rgba(255,255,255,0.2);stroke-width:1;stroke-dasharray:4,3}
.ell{font-size:9px;fill:#64748b;pointer-events:none;text-shadow:0 0 4px #0f172a}
.nc{cursor:pointer;transition:r 0.2s,opacity 0.2s}.nc:hover{filter:brightness(1.3)}
.nl{font-size:11px;fill:#e2e8f0;pointer-events:none;font-family:monospace;text-shadow:0 0 4px rgba(0,0,0,0.8)}
.nd{fill:#475569}.mk{fill:rgba(255,255,255,0.2)}
#st{position:fixed;top:96px;left:50%;transform:translateX(-50%);z-index:5;font-size:13px;color:#475569;pointer-events:none;text-align:center}
#dp{position:fixed;top:56px;right:0;bottom:0;width:460px;z-index:25;background:rgba(15,23,42,0.97);backdrop-filter:blur(20px);border-left:1px solid rgba(255,255,255,0.08);transform:translateX(100%);transition:transform 0.25s;overflow-y:auto;padding:24px}
#dp.o{transform:translateX(0)}#dc{position:sticky;top:0;float:right;cursor:pointer;font-size:24px;color:#64748b;padding:2px 8px;line-height:1}
.dt{font-size:16px;font-weight:600;color:#f1f5f9;word-break:break-all;margin-bottom:4px;padding-right:32px}
.dm{font-size:12px;color:#64748b;margin-bottom:16px;line-height:1.8}.dm span{display:inline-block;padding:1px 8px;border-radius:4px;font-size:11px;margin-left:6px}
.dd{color:#94a3b8;font-size:13px;margin-bottom:16px;padding:8px 12px;background:rgba(255,255,255,0.03);border-radius:6px}
.ci{font-size:12px;font-family:monospace;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.03);line-height:1.7}
.ci .c1{color:#64748b}.ci .c2{color:#e2e8f0}.ci .c3{color:#a5d6ff}.ci .c4{color:#475569;margin-left:8px;font-size:11px;font-family:sans-serif}
#ci{position:fixed;bottom:24px;right:24px;z-index:10;font-size:12px;color:#64748b;background:rgba(30,41,59,0.8);backdrop-filter:blur(8px);padding:8px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.04)}
</style>
</head>
<body>
<div id="hd"><h1>数据字典 · 关联模型</h1><span>'''+str(len(nodes))+'''表 · 搜索发现</span></div>
<div id="sr"><input type="text" id="si" placeholder="搜索表名 / 业务对象..." autocomplete="off"><div id="sg"></div></div>
<div id="st">输入关键字搜索</div>
<div id="lg"></div><div id="tp"></div>
<div id="dp"><div id="dc" onclick="cl()">&times;</div><div id="dc2"></div></div>
<div id="ci">99 表</div>
<svg id="g"></svg>
<script>
var IDX='''+j_idx+''';var NM='''+j_nm+''';var EG='''+j_eg+''';var CL='''+j_cl+''';var CR='''+j_cr+''';
''')
    
    # Now write the JavaScript logic piece by piece
    parts.append('''
var lg=document.getElementById("lg");
var ds=[...new Set(IDX.map(function(d){return d.domain;}))];
lg.innerHTML=ds.map(function(d){return"<span class=\"li\"><span class=\"ld\" style=\"background:"+(CR[d]||"#94a3b8")+"\"></span>"+d+"</span>";}).join("");
''')

    parts.append('''
var si=document.getElementById("si");
var sg=document.getElementById("sg");
var st=document.getElementById("st");
var tp=document.getElementById("tp");
var tt=null;

si.addEventListener("input",function(){
    clearTimeout(tt);
    var v=this.value.trim();
    if(!v){sg.style.display="none";st.textContent="输入关键字搜索";return;}
    tt=setTimeout(function(){
        var q=v.toLowerCase();
        var m=IDX.filter(function(d){return d.id.toLowerCase().indexOf(q)>=0||d.label.toLowerCase().indexOf(q)>=0||d.bo.toLowerCase().indexOf(q)>=0||d.desc.toLowerCase().indexOf(q)>=0;}).slice(0,30);
        if(m.length===0){sg.style.display="none";return;}
        sg.style.display="block";
        sg.innerHTML=m.map(function(d){
            var c=CR[d.domain]||"#94a3b8";
            return"<div class=\"si\" data-i=\""+d.id+"\"><span class=\"sl\">"+d.label+"</span><span class=\"sb\">"+d.bo+"</span><span class=\"sd\" style=\"background:"+c+"22;color:"+c+"\">"+d.domain+"</span></div>";
        }).join("");
        sg.querySelectorAll(".si").forEach(function(el){
            el.addEventListener("click",function(){search(this.getAttribute("data-i"));sg.style.display="none";});
        });
    },150);
});

function search(q){
    si.value=q;
    var ql=q.toLowerCase();
    var ms=new Set();
    IDX.forEach(function(d){if(d.id.toLowerCase().indexOf(ql)>=0||d.label.toLowerCase().indexOf(ql)>=0||d.bo.toLowerCase().indexOf(ql)>=0||d.desc.toLowerCase().indexOf(ql)>=0)ms.add(d.id);});
    if(ms.size===0){st.textContent="无匹配";return;}
    var vs=new Set(ms);
    EG.forEach(function(e){
        var s=e.source.id||e.source;
        var t=e.target.id||e.target;
        if(ms.has(s))vs.add(t);if(ms.has(t))vs.add(s);
    });
    draw(ms,vs);
    document.getElementById("ci").textContent=vs.size+" 表 \u00b7 "+ms.size+" 匹配";
}
''')

    parts.append('''
function draw(ms,vs){
    st.textContent="";
    document.getElementById("g").innerHTML="";
    var vn=[];vs.forEach(function(id){if(NM[id])vn.push(JSON.parse(JSON.stringify(NM[id])));});
    var ve=EG.filter(function(e){
        var s=e.source.id||e.source;var t=e.target.id||e.target;
        return vs.has(s)&&vs.has(t);
    });
    ve.forEach(function(e,i){e._idx=i;});
    
    var w=window.innerWidth,h=window.innerHeight;
    var svg=d3.select("#g").attr("viewBox",[0,0,w,h]);
    var g=svg.append("g");
    svg.call(d3.zoom().scaleExtent([0.15,4]).on("zoom",function(e){g.attr("transform",e.transform);})).on("dblclick.zoom",null);
    
    var df=svg.append("defs");
    ve.forEach(function(e,i){
        df.append("marker").attr("id","a"+i).attr("viewBox","0 -5 10 10").attr("refX",20).attr("refY",0).attr("markerWidth",6).attr("markerHeight",6).attr("orient","auto").append("path").attr("d","M0,-5L10,0L0,5").attr("class","mk");
    });
    
    var sim=d3.forceSimulation(vn).force("link",d3.forceLink(ve).id(function(d){return d.id;}).distance(function(d){return d.type==="cross"?200:130;})).force("charge",d3.forceManyBody().strength(-350)).force("center",d3.forceCenter(w/2,h/2)).force("collision",d3.forceCollide(35));
    var lk=g.append("g").selectAll("line").data(ve).join("line").attr("class",function(d){return d.type==="cross"?"el ec":"el";}).attr("marker-end",function(d){return"url(#a"+d._idx+")";});
    var ll=g.append("g").selectAll("text").data(ve).join("text").attr("class","ell").text(function(d){return d.label.substring(0,14);});
    var ng=g.append("g").selectAll("g").data(vn).join("g").call(d3.drag().on("start",function(e,d){if(!e.active)sim.alphaTarget(0.3).restart();d.fx=d.x;d.fy=d.y;}).on("drag",function(e,d){d.fx=e.x;d.fy=e.y;}).on("end",function(e,d){if(!e.active)sim.alphaTarget(0);d.fx=null;d.fy=null;}));
    
    ng.append("circle").attr("class","nc").attr("r",function(d){return d.type==="main"?11:7;}).attr("fill",function(d){return CR[d.domain]||"#94a3b8";}).attr("stroke","rgba(255,255,255,0.15)").attr("stroke-width",1.5);
    ng.append("text").attr("class","nl").attr("dx",function(d){return d.type==="main"?15:12;}).attr("dy",4).text(function(d){return d.label;});
    
    ng.on("mouseenter",function(e,d){
        var c=CR[d.domain]||"#94a3b8";
        tp.innerHTML="<div class=\"tt\">"+d.id+"</div><div class=\"tb\">"+d.bo+"</div>";
        var x=e.pageX+16,y=e.pageY-40;
        if(x+400>window.innerWidth)x=e.pageX-416;if(y<60)y=60;
        tp.style.left=x+"px";tp.style.top=y+"px";tp.classList.add("show");
        var cn=new Set();
        ve.forEach(function(ed){
            var s=ed.source.id||ed.source;var t=ed.target.id||ed.target;
            if(s===d.id||t===d.id){cn.add(s);cn.add(t);}
        });
        ng.selectAll("circle").attr("opacity",function(n){return cn.has(n.id)||n.id===d.id?1:0.08;});
        ng.selectAll("text").attr("class",function(n){return cn.has(n.id)||n.id===d.id?"nl":"nl nd";});
        lk.style("opacity",function(ed){
            var s=ed.source.id||ed.source;var t=ed.target.id||ed.target;
            return(s===d.id||t===d.id)?1:0.04;
        });
    }).on("mouseleave",function(){
        tp.classList.remove("show");
        ng.selectAll("circle").attr("opacity",1);
        ng.selectAll("text").attr("class","nl");
        lk.style("opacity",1);
    }).on("click",function(e,d){sd(d);});
    
    sim.on("tick",function(){
        lk.attr("x1",function(d){return d.source.x;}).attr("y1",function(d){return d.source.y;}).attr("x2",function(d){return d.target.x;}).attr("y2",function(d){return d.target.y;});
        ll.attr("x",function(d){return(d.source.x+d.target.x)/2;}).attr("y",function(d){return(d.source.y+d.target.y)/2-6;});
        ng.attr("transform",function(d){return"translate("+d.x+","+d.y+")";});
    });
}
''')

    parts.append('''
function cl(){document.getElementById("dp").classList.remove("o");}
function sd(d){
    var c=CR[d.domain]||"#94a3b8";
    var co=CL[d.id]||[];
    var h="<div class=\"dt\">"+d.id+"</div><div class=\"dm\">"+d.bo+" <span style=\"background:"+c+"22;color:"+c+";border:1px solid "+c+"44\">"+d.domain+"</span></div><div class=\"dd\">"+d.desc+"</div>";
    if(co.length>0){
        h+="<div style=\"font-size:12px;color:#64748b;margin-bottom:8px\">comment on table "+d.id+" is '<span style=\"color:#e2e8f0\">"+d.desc+"</span>';</div><div style=\"height:1px;background:rgba(255,255,255,0.06);margin:8px 0 12px 0\"></div>";
        co.forEach(function(o){
            var v=(o.r&&o.r!==o.n&&o.r!==o.t)?o.r:o.t;
            h+="<div class=\"ci\"><span class=\"c1\">comment on column </span><span class=\"c2\">"+d.id+"."+o.n+"</span><span class=\"c1\"> is '</span><span class=\"c3\">"+v+"</span><span class=\"c1\">';</span><span class=\"c4\">-- "+o.tp+"</span></div>";
        });
    }else{h+="<div style=\"color:#475569;font-size:13px;padding:24px 0;text-align:center\">暂无字段信息</div>";}
    document.getElementById("dc2").innerHTML=h;
    document.getElementById("dp").classList.add("o");
}

window.addEventListener("resize",function(){
    if(typeof draw==="undefined"||!window.sim)return;
    var w=window.innerWidth,h=window.innerHeight;
    d3.select("#g").attr("viewBox",[0,0,w,h]);
    sim.force("center",d3.forceCenter(w/2,h/2)).alpha(0.2).restart();
});
</script>
</body>
</html>''')

    with open(os.path.join(outdir,"graph.html"),"w",encoding="utf-8") as out:
        for part in parts:
            out.write(part)
    
    sz=os.path.getsize(os.path.join(outdir,"graph.html"))
    print("OK:",sz,"bytes -",len(nodes),"tables")
    print("Double-click graph.html to open")

if __name__=="__main__":
    main()
