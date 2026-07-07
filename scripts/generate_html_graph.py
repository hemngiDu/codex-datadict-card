import openpyxl, os, glob, json

cwd = os.getcwd()
import sys
fp = sys.argv[1] if len(sys.argv) > 1 else None
if not fp or not os.path.exists(fp):
    print("用法: python script.py <Excel文件路径>"); sys.exit(1)
wb = openpyxl.load_workbook(fp)
ws = wb["表目录"]

rows = []
current_no, current_bo = None, None
for r in range(7, ws.max_row + 1):
    # 表目录: col1~8, but first 3 are empty. col4=编号, col5=业务对象, col6=表名, col7=数据表名, col8=关联表
    c4 = ws.cell(r, 4).value  # 编号
    c5 = ws.cell(r, 5).value  # 业务对象
    c6 = ws.cell(r, 6).value  # 表名
    c7 = ws.cell(r, 7).value  # 数据表名
    c8 = ws.cell(r, 8).value  # 关联表

    if c4: 
        try: current_no = int(str(c4).strip())
        except: pass
    if c5: current_bo = str(c5).strip()
    if c6 and c7:
        rows.append({
            "no": current_no, "bo": current_bo,
            "desc": str(c6).strip(), "tbl": str(c7).strip(),
            "rel": str(c8).strip() if c8 else ""
        })

# Test
for row in rows[:3]:
    print("  bo=<<{}>> desc=<<{}>> tbl=<<{}>> rel=<<{}>>".format(
        row["bo"], row["desc"], row["tbl"], row["rel"]))

domain_map = ["配置与方案","因子与水位","算法与计算","执行与结果","日志审计","辅助测试"]
bo_to_domain = {}
for bo in ["库存计划方案","安全库存统计方案","日均消耗统计方案","调度方案"]: bo_to_domain[bo]=domain_map[0]
for bo in ["库存水位信息","库存水位因子","因子数据维护","因子取数方案","因子取数规则","补货日期"]: bo_to_domain[bo]=domain_map[1]
for bo in ["算法方案配置","算法注册配置","库存水位更新方案","智能计算设置","因子分批计算参数","快速搭建模型","资源注册模型"]: bo_to_domain[bo]=domain_map[2]
for bo in ["库存计划建议","供需匹配明细","安全库存记录","匹配映射配置","日均消耗记录","客户服务水平"]: bo_to_domain[bo]=domain_map[3]
for bo in ["消耗量计算日志","安全库存计算日志","库存计划运算日志"]: bo_to_domain[bo]=domain_map[4]
for bo in ["组织测试"]: bo_to_domain[bo]=domain_map[5]

nodes, edges, node_ids = [], [], set()
for row in rows:
    if row["tbl"] in node_ids: continue
    node_ids.add(row["tbl"])
    domain = bo_to_domain.get(row["bo"],"其他")
    node_type = "main" if not row["rel"] else "child"
    nodes.append({"id":row["tbl"],"label":row["tbl"].replace("t_invp_","").replace("tk_",""),"bo":row["bo"],"desc":row["desc"],"domain":domain,"type":node_type})

for row in rows:
    if row["rel"] and row["rel"] != row["tbl"] and row["tbl"] in node_ids and row["rel"] in node_ids:
        edges.append({"source":row["rel"],"target":row["tbl"],"label":row["desc"]})

cross_refs = [
    ("t_invp_scheme","t_invp_model_register","需求/供应来源模型"),
    ("t_invp_scheme","t_invp_algoconfig","算法方案"),
    ("t_invp_scheme","t_invp_calrulecfg","水位更新方案"),
    ("t_invp_scheme","t_invp_invlevel","库存水位信息"),
    ("t_invp_scheme","t_invp_matchconfig","匹配映射配置"),
    ("t_invp_planadvice","t_invp_scheme","所属计划方案"),
    ("t_invp_invlevel","t_invp_levelfactor","水位因子"),
    ("t_invp_matchdetail","t_invp_scheme","计划方案"),
    ("t_invp_smartcalccfg","t_invp_invlevel","库存水位"),
    ("t_invp_smartcalccfg","t_invp_calrulecfg","因子计算规则"),
    ("t_invp_smartcalccfg","t_invp_queryschema","因子取数方案"),
    ("t_invp_invleveldata","t_invp_levelfactor","水位因子"),
    ("t_invp_ssrecord","t_invp_ssdayscheme","安全库存方案"),
    ("t_invp_dac_record","t_invp_dailycomsumption","日均消耗方案"),
    ("t_invp_safestock_callog","t_invp_ssdayscheme","安全库存方案"),
]
for src, tgt, lbl in cross_refs:
    if src in node_ids and tgt in node_ids:
        edges.append({"source":src,"target":tgt,"label":lbl,"type":"cross"})

col_counts = {}
for sn in wb.sheetnames:
    if sn == "表目录": continue
    sws = wb[sn]
    for r in range(1, sws.max_row+1):
        c2 = sws.cell(r,2).value
        c5 = sws.cell(r,5).value
        if c2 and str(c2).strip() == "数据表名：" and c5:
            tbl_name = str(c5).strip()
            if tbl_name not in col_counts: col_counts[tbl_name] = 0
            for rr in range(r+2, sws.max_row+1):
                cc2 = sws.cell(rr,3).value
                cc1 = sws.cell(rr,2).value
                if cc2 and cc1:
                    try: float(str(cc1).strip()); col_counts[tbl_name] = col_counts.get(tbl_name,0) + 1
                    except: break
                else: break
for n in nodes:
    n["colCount"] = col_counts.get(n["id"], 0)

print("\nNodes:", len(nodes))
print("Edges (from 关联表):", sum(1 for e in edges if "type" not in e))
print("Edges (cross-ref):", sum(1 for e in edges if e.get("type")=="cross"))
print("Total edges:", len(edges))

domain_colors = {"配置与方案":"#3b82f6","因子与水位":"#10b981","算法与计算":"#f59e0b","执行与结果":"#ef4444","日志审计":"#6b7280","辅助测试":"#8b5cf6","其他":"#94a3b8"}

data_json = json.dumps({"nodes":nodes,"edges":edges}, ensure_ascii=False)
colors_json = json.dumps(domain_colors, ensure_ascii=False)

html = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>供应链云 · 数据字典关联模型</title>
<script src="https://d3js.org/d3.v7.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0f172a;color:#e2e8f0;overflow:hidden}
#search{position:fixed;top:54px;left:50%;transform:translateX(-50%);z-index:15;width:340px}
#search input{width:100%;padding:8px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.08);background:rgba(30,41,59,0.9);backdrop-filter:blur(12px);color:#e2e8f0;font-size:13px;outline:none;transition:border-color 0.2s;box-sizing:border-box}
#search input:focus{border-color:rgba(255,255,255,0.25)}
#search input::placeholder{color:#475569}
.edge-line.search-dim{stroke:rgba(255,255,255,0.02)}
#header{position:fixed;top:0;left:0;right:0;z-index:10;padding:12px 24px;background:rgba(15,23,42,0.9);backdrop-filter:blur(12px);border-bottom:1px solid rgba(255,255,255,0.06);display:flex;align-items:center;gap:12px}
#header h1{font-size:18px;font-weight:600}
#header span{font-size:13px;color:#94a3b8}
#legend{position:fixed;bottom:24px;left:24px;z-index:10;display:flex;gap:8px;flex-wrap:wrap;background:rgba(30,41,59,0.9);backdrop-filter:blur(12px);padding:10px 14px;border-radius:10px;border:1px solid rgba(255,255,255,0.06)}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:#cbd5e1}
.legend-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0}
#tooltip{position:fixed;z-index:20;background:rgba(30,41,59,0.96);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:14px 18px;font-size:13px;line-height:1.6;max-width:400px;pointer-events:none;opacity:0;transition:opacity 0.15s;box-shadow:0 8px 32px rgba(0,0,0,0.4)}
#tooltip.show{opacity:1}
#tooltip .tt-title{font-size:15px;font-weight:600;color:#f1f5f9;margin-bottom:4px;word-break:break-all}
#tooltip .tt-bo{color:#94a3b8;font-size:12px;margin-bottom:6px}
#tooltip .tt-desc{color:#cbd5e1;margin-bottom:4px}
#tooltip .tt-col{color:#64748b;font-size:12px}
#tooltip .tt-domain{display:inline-block;padding:1px 8px;border-radius:4px;font-size:11px;margin-top:6px}
svg{position:fixed;top:0;left:0;width:100%;height:100%}
.edge-line{stroke:rgba(255,255,255,0.12);stroke-width:1.5;fill:none;transition:stroke 0.2s,stroke-width 0.2s}
.edge-line.highlight{stroke:rgba(255,255,255,0.5);stroke-width:2.5}
.edge-cross{stroke:rgba(255,255,255,0.18);stroke-width:1;stroke-dasharray:4,3}
.edge-label{font-size:9px;fill:#64748b;pointer-events:none;text-shadow:0 0 4px #0f172a}
.node-circle{cursor:pointer;transition:r 0.2s,opacity 0.2s}
.node-circle:hover{filter:brightness(1.3)}
.node-label{font-size:11px;fill:#e2e8f0;pointer-events:none;font-family:monospace;text-shadow:0 0 4px rgba(0,0,0,0.8)}
.node-label.dim{fill:#475569}
.marker{fill:rgba(255,255,255,0.2)}
#title-info{position:fixed;top:64px;left:50%;transform:translateX(-50%);z-index:5;font-size:14px;color:#94a3b8;pointer-events:none;text-align:center}
#count-info{position:fixed;bottom:24px;right:24px;z-index:10;font-size:12px;color:#94a3b8;background:rgba(30,41,59,0.8);backdrop-filter:blur(8px);padding:8px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.04)}
</style>
</head>
<body>
<div id="header"><h1>供应链云 · 数据字典</h1><span>库存计划模块 · """ + str(len(nodes)) + """张表 · """ + str(len(edges)) + """条关联</span></div>
<div id="title-info">拖拽节点 · 滚轮缩放 · 悬停高亮关联路径</div>
<div id="legend"></div>
<div id="search"><input type="text" id="searchInput" placeholder="搜索表名 / 业务对象..." oninput="filterNodes(this.value)"></div>
<div id="tooltip"></div>
<div id="count-info">实线=父子表 · 虚线=跨对象引用</div>
<svg id="graph"></svg>
<script>
const rawData = """ + data_json + r""";
const nodes = rawData.nodes.map(n=>({...n}));
const edges = rawData.edges;
const domainColors = """ + colors_json + r""";

const legendEl = document.getElementById("legend");
[...new Set(nodes.map(n=>n.domain))].forEach(d=>{
    const item=document.createElement("div");
    item.className="legend-item";
    item.innerHTML="<span class=\"legend-dot\" style=\"background:"+(domainColors[d]||"#94a3b8")+"\"></span>"+d;
    legendEl.appendChild(item);
});

const svg = d3.select("#graph");
const width=window.innerWidth, height=window.innerHeight;
svg.attr("viewBox",[0,0,width,height]);
const g=svg.append("g");

svg.call(d3.zoom().scaleExtent([0.15,4]).on("zoom",e=>g.attr("transform",e.transform))).on("dblclick.zoom",null);

const sim=d3.forceSimulation(nodes)
    .force("link",d3.forceLink(edges).id(d=>d.id).distance(d=>d.type==="cross"?200:130))
    .force("charge",d3.forceManyBody().strength(-350))
    .force("center",d3.forceCenter(width/2,height/2))
    .force("collision",d3.forceCollide(35));

const defs=svg.append("defs");
edges.forEach((e,i)=>{
    defs.append("marker").attr("id","a"+i).attr("viewBox","0 -5 10 10")
        .attr("refX",20).attr("refY",0).attr("markerWidth",6).attr("markerHeight",6).attr("orient","auto")
        .append("path").attr("d","M0,-5L10,0L0,5").attr("class","marker");
});

const link=g.append("g").selectAll("line").data(edges).join("line")
    .attr("class",d=>d.type==="cross"?"edge-line edge-cross":"edge-line")
    .attr("marker-end",(d,i)=>"url(#a"+i+")");

const linkLabel=g.append("g").selectAll("text").data(edges).join("text")
    .attr("class","edge-label").text(d=>d.label.substring(0,14));

const nodeG=g.append("g").selectAll("g").data(nodes).join("g")
    .call(d3.drag().on("start",(e,d)=>{if(!e.active)sim.alphaTarget(0.3).restart();d.fx=d.x;d.fy=d.y;})
        .on("drag",(e,d)=>{d.fx=e.x;d.fy=e.y;})
        .on("end",(e,d)=>{if(!e.active)sim.alphaTarget(0);d.fx=null;d.fy=null;}));

nodeG.append("circle").attr("class","node-circle")
    .attr("r",d=>d.type==="main"?11:7)
    .attr("fill",d=>domainColors[d.domain]||"#94a3b8")
    .attr("stroke","rgba(255,255,255,0.15)").attr("stroke-width",1.5);

nodeG.append("text").attr("class","node-label")
    .attr("dx",d=>d.type==="main"?15:12).attr("dy",4)
    .text(d=>d.label);

const tip=document.getElementById("tooltip");
function show(e,d){
    const c=domainColors[d.domain]||"#94a3b8";
    tip.innerHTML="<div class=\"tt-title\">"+d.id+"</div><div class=\"tt-bo\">"+d.bo+"</div><div class=\"tt-desc\">"+d.desc+"</div><div class=\"tt-col\">"+d.colCount+"个字段</div><span class=\"tt-domain\" style=\"background:"+c+"22;color:"+c+";border:1px solid "+c+"44\">"+d.domain+"</span>";
    let x=e.pageX+16,y=e.pageY-40;
    if(x+400>window.innerWidth)x=e.pageX-416;
    if(y<60)y=60;
    tip.style.left=x+"px";tip.style.top=y+"px";tip.classList.add("show");
    const cn=new Set();
    edges.forEach(ed=>{if(ed.source.id===d.id||ed.target.id===d.id){cn.add(ed.source.id);cn.add(ed.target.id)}});
    nodeG.selectAll("circle").attr("opacity",n=>cn.has(n.id)||n.id===d.id?1:0.08);
    nodeG.selectAll("text").attr("class",n=>cn.has(n.id)||n.id===d.id?"node-label":"node-label dim");
    link.attr("class",ed=>(ed.source.id===d.id||ed.target.id===d.id)?(ed.type==="cross"?"edge-line edge-cross highlight":"edge-line highlight"):(ed.type==="cross"?"edge-line edge-cross":"edge-line")).style("opacity",ed=>(ed.source.id===d.id||ed.target.id===d.id)?1:0.04);
}
function hide(){
    tip.classList.remove("show");
    nodeG.selectAll("circle").attr("opacity",1);
    nodeG.selectAll("text").attr("class","node-label");
    link.attr("class",d=>d.type==="cross"?"edge-line edge-cross":"edge-line").style("opacity",1);
}
nodeG.on("mouseenter",show).on("mousemove",show).on("mouseleave",hide);

sim.on("tick",()=>{
    link.attr("x1",d=>d.source.x).attr("y1",d=>d.source.y).attr("x2",d=>d.target.x).attr("y2",d=>d.target.y);
    linkLabel.attr("x",d=>(d.source.x+d.target.x)/2).attr("y",d=>(d.source.y+d.target.y)/2-6);
    nodeG.attr("transform",d=>"translate("+d.x+","+d.y+")");
});


function filterNodes(query){
    var q=query.trim().toLowerCase();
    if(!q){
        nodeG.selectAll("circle").attr("opacity",1);
        nodeG.selectAll("text").attr("class","node-label");
        link.attr("class",function(d){return d.type==="cross"?"edge-line edge-cross":"edge-line";}).style("opacity",1);
        document.getElementById("searchInput").placeholder="搜索表名 / 业务对象...";
        return;
    }
    var matchIds=new Set();
    nodes.forEach(function(n){
        if(n.label.toLowerCase().includes(q)||(n.bo||"").toLowerCase().includes(q)||(n.desc||"").toLowerCase().includes(q)){
            matchIds.add(n.id);
        }
    });
    var visibleIds=new Set(matchIds);
    edges.forEach(function(e){
        var srcId=e.source.id||e.source;
        var tgtId=e.target.id||e.target;
        if(matchIds.has(srcId))visibleIds.add(tgtId);
        if(matchIds.has(tgtId))visibleIds.add(srcId);
    });
    nodeG.selectAll("circle").attr("opacity",function(d){
        return visibleIds.has(d.id)?1:0.03;
    });
    nodeG.selectAll("text").attr("class",function(d){
        return visibleIds.has(d.id)?"node-label":"node-label dim";
    });
    link.attr("class",function(d){
        var srcId=d.source.id||d.source;
        var tgtId=d.target.id||d.target;
        var show=visibleIds.has(srcId)&&visibleIds.has(tgtId);
        return (d.type==="cross"?"edge-line edge-cross":"edge-line")+(show?"":" search-dim");
    }).style("opacity",function(d){
        var srcId=d.source.id||d.source;
        var tgtId=d.target.id||d.target;
        return visibleIds.has(srcId)&&visibleIds.has(tgtId)?1:0.03;
    });
    document.getElementById("searchInput").placeholder=matchIds.size+" 条匹配 \u00b7 "+visibleIds.size+" 关联节点";
}

window.addEventListener("resize",()=>{
    const w=window.innerWidth,h=window.innerHeight;
    svg.attr("viewBox",[0,0,w,h]);
    sim.force("center",d3.forceCenter(w/2,h/2)).alpha(0.2).restart();
});
</script>
</body>
</html>
"""

outpath = os.path.join(os.path.dirname(os.path.abspath(fp)), os.path.splitext(os.path.basename(fp))[0] + "_graph.html")
with open(outpath, "w", encoding="utf-8") as f:
    f.write(html)
print("\nGenerated:", outpath)