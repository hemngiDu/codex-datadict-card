#!/usr/bin/env python3
"""Scalable graph: graph.html + data.js, works with file:// (no server needed)"""
import openpyxl, os, sys, json, glob

DOMAIN_GROUPS = [
    ("配置与方案",["库存计划方案","安全库存统计方案","日均消耗统计方案","调度方案"]),
    ("因子与水位",["库存水位信息","库存水位因子","因子数据维护","因子取数方案","因子取数规则","补货日期"]),
    ("算法与计算",["算法方案配置","算法注册配置","库存水位更新方案","智能计算设置","因子分批计算参数","快速搭建模型","资源注册模型"]),
    ("执行与结果",["库存计划建议","供需匹配明细","安全库存记录","匹配映射配置","日均消耗记录","客户服务水平"]),
    ("日志审计",["消耗量计算日志","安全库存计算日志","库存计划运算日志"]),
    ("辅助测试",["组织测试"]),
]
CROSS_REFS = [
    ("t_invp_scheme","t_invp_model_register","需求/供应来源模型"),
    ("t_invp_scheme","t_invp_algoconfig","算法方案"),
    ("t_invp_scheme","t_invp_calrulecfg","水位更新方案"),
    ("t_invp_scheme","t_invp_invlevel","库存水位信息"),
    ("t_invp_scheme","t_invp_matchconfig","匹配映射配置"),
    ("t_invp_planadvice","t_invp_scheme","所属计划方案"),
    ("t_invp_invlevel","t_invp_levelfactor","水位因子"),
    ("t_invp_matchdetail","t_invp_scheme","计划方案"),
    ("t_invp_smartcalccfg","t_invp_invlevel","库存水位"),
    ("t_invp_smartcalccfg","t_invp_calrulecfg","因子计算规则"),
    ("t_invp_smartcalccfg","t_invp_queryschema","因子取数方案"),
    ("t_invp_invleveldata","t_invp_levelfactor","水位因子"),
    ("t_invp_ssrecord","t_invp_ssdayscheme","安全库存方案"),
    ("t_invp_dac_record","t_invp_dailycomsumption","日均消耗方案"),
    ("t_invp_safestock_callog","t_invp_ssdayscheme","安全库存方案"),
]
DOMAIN_COLORS = {"配置与方案":"#3b82f6","因子与水位":"#10b981","算法与计算":"#f59e0b","执行与结果":"#ef4444","日志审计":"#6b7280","辅助测试":"#8b5cf6","其他":"#94a3b8"}

def get_domain(bo):
    for d, objs in DOMAIN_GROUPS: return d if bo in objs else None
    return "其他"
def get_domain_color(domain):
    return DOMAIN_COLORS.get(domain, "#94a3b8")

def parse_catalog(wb):
    ws = wb["表目录"]; rows = []; curr_no = curr_bo = None
    for r in range(7, ws.max_row + 1):
        c4=ws.cell(r,4).value;c5=ws.cell(r,5).value;c6=ws.cell(r,6).value;c7=ws.cell(r,7).value;c8=ws.cell(r,8).value
        if c4:
            try: curr_no=int(str(c4).strip())
            except: pass
        if c5: curr_bo=str(c5).strip()
        if c6 and c7: rows.append({"no":curr_no,"bo":curr_bo,"desc":str(c6).strip(),"tbl":str(c7).strip(),"rel":str(c8).strip() if c8 else ""})
    return rows

def get_column_details(wb):
    details = {}
    for sn in wb.sheetnames:
        if sn=="表目录": continue
        sws=wb[sn];cur_tbl=None;cols=[];in_hdr=False
        for r in range(1,sws.max_row+1):
            c2=sws.cell(r,2).value;c5=sws.cell(r,5).value;c8=sws.cell(r,8).value
            if c2 and str(c2).strip()=="数据表名：" and c5:
                if cur_tbl and cols:
                    if cur_tbl not in details or len(cols)>len(details.get(cur_tbl,[])): details[cur_tbl]=cols
                cur_tbl=str(c5).strip();cols=[];in_hdr=False
            elif c2 and str(c2).strip()=="编号":
                c3=sws.cell(r,3).value;c4=sws.cell(r,4).value
                if c3 and c4 and str(c3).strip()=="列名" and "列标题"in str(c4): in_hdr=True
            elif in_hdr and c2:
                try:
                    float(str(c2).strip())
                    cn=str(sws.cell(r,3).value).strip() if sws.cell(r,3).value else ""
                    ct=str(sws.cell(r,4).value).strip() if sws.cell(r,4).value else cn
                    tp=str(sws.cell(r,5).value).strip() if sws.cell(r,5).value else ""
                    rr=str(sws.cell(r,11).value).strip() if sws.cell(r,11).value else ""
                    if cn: cols.append({"n":cn,"t":ct,"tp":tp,"r":rr})
                except:
                    if c8 and str(c8).strip()=="返回主目录": in_hdr=False
        if cur_tbl and cols:
            if cur_tbl not in details or len(cols)>len(details.get(cur_tbl,[])): details[cur_tbl]=cols
    return details

def main():
    if len(sys.argv)<2:
        print("Usage: python generate.py <Excel file or dir>"); sys.exit(1)
    path=sys.argv[1]
    if os.path.isdir(path):
        files=[f for f in os.listdir(path) if f.endswith(".xlsx") and not f.startswith("~$")]
        if not files: print("No xlsx found"); sys.exit(1)
        path=os.path.join(path,files[0])
    base=os.path.splitext(os.path.basename(path))[0]
    outdir=os.path.dirname(os.path.abspath(path))
    print("Reading:", path)
    wb=openpyxl.load_workbook(path)
    rows=parse_catalog(wb)
    print("Tables:", len(rows))

    # Build nodes, edges, index
    nids=set();nodes=[];edges=[]
    for row in rows:
        if row["tbl"] in nids: continue
        nids.add(row["tbl"]); d=get_domain(row["bo"])
        nt="main" if not row["rel"] else "child"
        nodes.append({"id":row["tbl"],"label":row["tbl"].replace("t_invp_","").replace("tk_",""),"bo":row["bo"],"desc":row["desc"],"domain":d,"type":nt})
    for row in rows:
        if row["rel"] and row["rel"]!=row["tbl"] and row["tbl"] in nids and row["rel"] in nids:
            edges.append({"source":row["rel"],"target":row["tbl"],"label":row["desc"]})
    for src,tgt,lbl in CROSS_REFS:
        if src in nids and tgt in nids: edges.append({"source":src,"target":tgt,"label":lbl,"type":"cross"})

    bo_tables={}
    for n in nodes:
        bo=n["bo"]
        bo_tables.setdefault(bo,[]).append(n["id"])

    col_details=get_column_details(wb)

    # Generate data.js - all data as JS variables
    js_lines=['var INDEX_DATA = '+json.dumps([{"id":n["id"],"label":n["label"],"bo":n["bo"],"domain":n["domain"],"desc":n["desc"],"chunk":n["bo"].replace(" ","")} for n in nodes],ensure_ascii=False)+';']
    js_lines.append('var NODE_DATA = '+json.dumps({n["id"]:n for n in nodes},ensure_ascii=False)+';')
    js_lines.append('var EDGE_DATA = '+json.dumps(edges,ensure_ascii=False)+';')
    js_lines.append('var COL_DATA = '+json.dumps(col_details,ensure_ascii=False,default=lambda x:str(x))+';')
    data_js = '\n'.join(js_lines)

    index_json = json.dumps([{"id":n["id"],"label":n["label"],"bo":n["bo"],"domain":n["domain"],"desc":n["desc"],"chunk":n["bo"].replace(" ","")} for n in nodes],ensure_ascii=False)
    nodes_json = json.dumps({n["id"]:n for n in nodes},ensure_ascii=False)
    edges_json = json.dumps(edges,ensure_ascii=False)
    cols_json_all = json.dumps(col_details,ensure_ascii=False,default=lambda x:str(x))
    colors_json = json.dumps(DOMAIN_COLORS,ensure_ascii=False)
    xref_json = json.dumps([e for e in edges if e.get("type")=="cross"],ensure_ascii=False)

    # Generate graph.html with inline index + data.js loader
    html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>数据字典 · 关联模型</title>
<script src="https://d3js.org/d3.v7.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0f172a;color:#e2e8f0;overflow:hidden}
#header{position:fixed;top:0;left:0;right:0;z-index:10;padding:12px 24px;background:rgba(15,23,42,0.9);backdrop-filter:blur(12px);border-bottom:1px solid rgba(255,255,255,0.06);display:flex;align-items:center;gap:12px}
#header h1{font-size:18px;font-weight:600}#header span{font-size:13px;color:#94a3b8}
#search-area{position:fixed;top:56px;left:50%;transform:translateX(-50%);z-index:15;width:420px;padding:10px 0}
#search-area input{width:100%;padding:9px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.08);background:rgba(30,41,59,0.95);backdrop-filter:blur(12px);color:#e2e8f0;font-size:13px;outline:none;transition:border-color 0.2s;box-sizing:border-box}
#search-area input:focus{border-color:rgba(255,255,255,0.25)}#search-area input::placeholder{color:#475569}
#suggestions{position:absolute;top:48px;left:0;right:0;background:rgba(30,41,59,0.98);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.08);border-radius:8px;max-height:320px;overflow-y:auto;display:none}
.sg-item{padding:8px 14px;cursor:pointer;font-size:13px;border-bottom:1px solid rgba(255,255,255,0.03);display:flex;align-items:center;gap:8px;transition:background 0.15s}
.sg-item:hover{background:rgba(255,255,255,0.06)}.sg-item .sg-label{font-family:monospace;color:#e2e8f0}
.sg-item .sg-bo{color:#64748b;font-size:11px}
.sg-item .sg-domain{font-size:10px;padding:1px 6px;border-radius:3px;margin-left:auto}
#legend{position:fixed;bottom:24px;left:24px;z-index:10;display:flex;gap:8px;flex-wrap:wrap;background:rgba(30,41,59,0.9);backdrop-filter:blur(12px);padding:10px 14px;border-radius:10px;border:1px solid rgba(255,255,255,0.06)}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:#cbd5e1}
.legend-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0}
#tooltip{position:fixed;z-index:20;background:rgba(30,41,59,0.96);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:14px 18px;font-size:13px;line-height:1.6;max-width:400px;pointer-events:none;opacity:0;transition:opacity 0.15s;box-shadow:0 8px 32px rgba(0,0,0,0.4)}
#tooltip.show{opacity:1}#tooltip .tt-title{font-size:15px;font-weight:600;color:#f1f5f9;word-break:break-all}
#tooltip .tt-bo{color:#94a3b8;font-size:12px;margin-bottom:6px}
svg{position:fixed;top:0;left:0;width:100%;height:100%}
.edge-line{stroke:rgba(255,255,255,0.12);stroke-width:1.5;fill:none;transition:stroke 0.2s}
.edge-cross{stroke:rgba(255,255,255,0.2);stroke-width:1;stroke-dasharray:4,3}
.edge-label{font-size:9px;fill:#64748b;pointer-events:none;text-shadow:0 0 4px #0f172a}
.node-circle{cursor:pointer;transition:r 0.2s,opacity 0.2s}.node-circle:hover{filter:brightness(1.3)}
.node-label{font-size:11px;fill:#e2e8f0;pointer-events:none;font-family:monospace;text-shadow:0 0 4px rgba(0,0,0,0.8)}
.node-label.dim{fill:#475569}.marker{fill:rgba(255,255,255,0.2)}
#status{position:fixed;top:96px;left:50%;transform:translateX(-50%);z-index:5;font-size:13px;color:#475569;pointer-events:none;text-align:center}
#detail-panel{position:fixed;top:56px;right:0;bottom:0;width:480px;z-index:25;background:rgba(15,23,42,0.97);backdrop-filter:blur(20px);border-left:1px solid rgba(255,255,255,0.08);transform:translateX(100%);transition:transform 0.25s cubic-bezier(0.4,0,0.2,1);overflow-y:auto;padding:24px}
#detail-panel.open{transform:translateX(0)}#detail-close{position:sticky;top:0;float:right;cursor:pointer;font-size:24px;color:#64748b;padding:2px 8px;border-radius:4px;line-height:1}
#detail-close:hover{color:#e2e8f0;background:rgba(255,255,255,0.06)}
.dp-title{font-size:16px;font-weight:600;color:#f1f5f9;word-break:break-all;margin-bottom:4px;padding-right:32px}
.dp-meta{font-size:12px;color:#64748b;margin-bottom:6px;line-height:1.8}
.dp-meta span{display:inline-block;padding:1px 8px;border-radius:4px;font-size:11px;margin-left:6px}
.dp-desc{color:#94a3b8;font-size:13px;margin-bottom:16px;padding:8px 12px;background:rgba(255,255,255,0.03);border-radius:6px}
.col-item{font-size:12px;font-family:monospace;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.03);line-height:1.7}
.col-item .kw{color:#64748b}.col-item .tbl{color:#e2e8f0}.col-item .val{color:#a5d6ff}
.col-item .type{color:#475569;margin-left:8px;font-size:11px;font-family:sans-serif}
#count-info{position:fixed;bottom:24px;right:24px;z-index:10;font-size:12px;color:#64748b;background:rgba(30,41,59,0.8);backdrop-filter:blur(8px);padding:8px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.04)}
</style>
</head>
<body>
<div id="header"><h1>数据字典 · 关联模型</h1><span>搜索发现 · 按需加载</span></div>
<div id="search-area"><input type="text" id="searchInput" placeholder="搜索表名 / 业务对象..." autocomplete="off"><div id="suggestions"></div></div>
<div id="status">加载数据中...</div>
<div id="legend"></div><div id="tooltip"></div>
<div id="detail-panel"><div id="detail-close" onclick="closeDetail()">&times;</div><div id="detail-content"></div></div>
<div id="count-info">0 表</div>
<svg id="graph"></svg>
<script>
// ====== Inline search index ======
var INDEX_DATA = ''' + index_json + ''';
var COLORS = ''' + colors_json + ''';
var XREF = ''' + xref_json + ''';

// ====== Bulk data (loaded from data.js) ======
var NODES = {};
var EDGES = [];
var COLS = {};
var dataLoaded = false;
var dataLoading = false;

function loadBulkData(){
    if(dataLoaded || dataLoading) return;
    dataLoading = true;
    document.getElementById("status").textContent = "正在加载数据("+INDEX_DATA.length+"表)...";
    var s = document.createElement("script");
    s.src = "data.js";
    s.onload = function(){
        dataLoaded = true;
        dataLoading = false;
        document.getElementById("status").textContent = "";
        document.getElementById("count-info").textContent = INDEX_DATA.length + " 表";
        var params = new URLSearchParams(window.location.search);
        var q = params.get("q");
        if(q){ document.getElementById("searchInput").value = q; doSearch(q); }
    };
    document.head.appendChild(s);
}
setTimeout(loadBulkData, 100);

// ====== Legend ======
var legendEl = document.getElementById("legend");
var domains = [...new Set(INDEX_DATA.map(function(d){return d.domain;}))];
legendEl.innerHTML = domains.map(function(d){
    return "<div class=\"legend-item\"><span class=\"legend-dot\" style=\"background:"+(COLORS[d]||"#94a3b8")+"\"></span>"+d+"</div>";
}).join("");

// ====== Search ======
var searchTimer = null;
document.getElementById("searchInput").addEventListener("input", function(){
    clearTimeout(searchTimer);
    var v = this.value.trim();
    if(!v){ document.getElementById("suggestions").style.display="none"; return; }
    searchTimer = setTimeout(function(){showSuggestions(v);}, 150);
});

function showSuggestions(q){
    var ql = q.toLowerCase();
    var matches = INDEX_DATA.filter(function(d){
        return d.id.toLowerCase().includes(ql) || d.label.toLowerCase().includes(ql) ||
               d.bo.toLowerCase().includes(ql) || d.desc.toLowerCase().includes(ql);
    }).slice(0, 30);
    var el = document.getElementById("suggestions");
    if(matches.length===0){ el.style.display="none"; return; }
    el.innerHTML = matches.map(function(d){
        var c = COLORS[d.domain]||"#94a3b8";
        return "<div class=\"sg-item\" onclick=\"doSearch(\'"+d.id.replace(/'/g,"")+"\')\"><span class=\"sg-label\">"+d.label+"</span><span class=\"sg-bo\">"+d.bo+"</span><span class=\"sg-domain\" style=\"background:"+c+"22;color:"+c+"\">"+d.domain+"</span></div>";
    }).join("");
    el.style.display = "block";
}

// ====== Graph rendering ======
var svg = null, g = null, sim = null, nodeG = null, link = null, linkLabel = null;

function doSearch(q){
    document.getElementById("suggestions").style.display="none";
    var ql = q.toLowerCase();
    var matchIds = new Set();
    INDEX_DATA.forEach(function(d){
        if(d.id.toLowerCase().includes(ql) || d.label.toLowerCase().includes(ql) || d.bo.toLowerCase().includes(ql) || d.desc.toLowerCase().includes(ql)){
            matchIds.add(d.id);
        }
    });
    if(matchIds.size===0){ document.getElementById("status").textContent="无匹配结果"; return; }
    
    // Find 1-hop neighbors from NODES/EDGES
    var visibleIds = new Set(matchIds);
    EDGES.forEach(function(e){
        var src = typeof e.source==="object"?e.source.id:e.source;
        var tgt = typeof e.target==="object"?e.target.id:e.target;
        if(matchIds.has(src)) visibleIds.add(tgt);
        if(matchIds.has(tgt)) visibleIds.add(src);
    });
    // Also check XREF
    XREF.forEach(function(e){
        if(matchIds.has(e.source)) visibleIds.add(e.target);
        if(matchIds.has(e.target)) visibleIds.add(e.source);
    });
    
    renderGraph(matchIds, visibleIds);
    document.getElementById("count-info").textContent = visibleIds.size + " 表 · " + matchIds.size + " 匹配";
}

function renderGraph(matchIds, visibleIds){
    document.getElementById("status").textContent = "";
    d3.select("svg").selectAll("*:not(defs)").remove();
    
    var visNodes = [];
    visibleIds.forEach(function(id){
        if(NODES[id]) visNodes.push(JSON.parse(JSON.stringify(NODES[id])));
    });
    
    var edgeList = EDGES.concat(XREF);
    var visEdges = edgeList.filter(function(e){
        var src = typeof e.source==="object"?e.source.id:e.source;
        var tgt = typeof e.target==="object"?e.target.id:e.target;
        return visibleIds.has(src) && visibleIds.has(tgt);
    });
    visEdges.forEach(function(e,i){ e._idx = i; });
    
    var width = window.innerWidth, height = window.innerHeight;
    svg = d3.select("#graph").attr("viewBox",[0,0,width,height]);
    g = svg.append("g");
    svg.call(d3.zoom().scaleExtent([0.15,4]).on("zoom",function(e){g.attr("transform",e.transform);})).on("dblclick.zoom",null);
    
    var defs = svg.append("defs");
    visEdges.forEach(function(e,i){
        defs.append("marker").attr("id","a"+i).attr("viewBox","0 -5 10 10")
            .attr("refX",20).attr("refY",0).attr("markerWidth",6).attr("markerHeight",6).attr("orient","auto")
            .append("path").attr("d","M0,-5L10,0L0,5").attr("class","marker");
    });
    
    sim = d3.forceSimulation(visNodes)
        .force("link",d3.forceLink(visEdges).id(function(d){return d.id;}).distance(function(d){return d.type==="cross"?200:130;}))
        .force("charge",d3.forceManyBody().strength(-350))
        .force("center",d3.forceCenter(width/2,height/2))
        .force("collision",d3.forceCollide(35));
    
    link = g.append("g").selectAll("line").data(visEdges).join("line")
        .attr("class",function(d){return d.type==="cross"?"edge-line edge-cross":"edge-line";})
        .attr("marker-end",function(d){return "url(#a"+d._idx+")";});
    
    linkLabel = g.append("g").selectAll("text").data(visEdges).join("text")
        .attr("class","edge-label").text(function(d){return d.label.substring(0,14);});
    
    nodeG = g.append("g").selectAll("g").data(visNodes).join("g")
        .call(d3.drag().on("start",function(e,d){if(!e.active)sim.alphaTarget(0.3).restart();d.fx=d.x;d.fy=d.y;})
            .on("drag",function(e,d){d.fx=e.x;d.fy=e.y;})
            .on("end",function(e,d){if(!e.active)sim.alphaTarget(0);d.fx=null;d.fy=null;}));
    
    nodeG.append("circle").attr("class","node-circle")
        .attr("r",function(d){return d.type==="main"?11:7;})
        .attr("fill",function(d){return COLORS[d.domain]||"#94a3b8";})
        .attr("stroke","rgba(255,255,255,0.15)").attr("stroke-width",1.5);
    
    nodeG.append("text").attr("class","node-label")
        .attr("dx",function(d){return d.type==="main"?15:12;}).attr("dy",4)
        .text(function(d){return d.label;});
    
    // Tooltip
    var tip = document.getElementById("tooltip");
    function showTt(e,d){
        var c=COLORS[d.domain]||"#94a3b8";
        tip.innerHTML="<div class=\"tt-title\">"+d.id+"</div><div class=\"tt-bo\">"+d.bo+"</div>";
        var x=e.pageX+16,y=e.pageY-40;
        if(x+400>window.innerWidth)x=e.pageX-416;if(y<60)y=60;
        tip.style.left=x+"px";tip.style.top=y+"px";tip.classList.add("show");
        var cn=new Set();
        visEdges.forEach(function(ed){
            var s=typeof ed.source==="object"?ed.source.id:ed.source;
            var t=typeof ed.target==="object"?ed.target.id:ed.target;
            if(s===d.id||t===d.id){cn.add(s);cn.add(t);}
        });
        nodeG.selectAll("circle").attr("opacity",function(n){return cn.has(n.id)||n.id===d.id?1:0.08;});
        nodeG.selectAll("text").attr("class",function(n){return cn.has(n.id)||n.id===d.id?"node-label":"node-label dim";});
        link.style("opacity",function(ed){
            var s=typeof ed.source==="object"?ed.source.id:ed.source;
            var t=typeof ed.target==="object"?ed.target.id:ed.target;
            return(s===d.id||t===d.id)?1:0.04;
        });
    }
    function hideTt(){
        tip.classList.remove("show");
        nodeG.selectAll("circle").attr("opacity",1);
        nodeG.selectAll("text").attr("class","node-label");
        link.style("opacity",1);
    }
    nodeG.on("mouseenter",showTt).on("mousemove",showTt).on("mouseleave",hideTt);
    nodeG.on("click",function(e,d){showDetail(d);});
    
    sim.on("tick",function(){
        link.attr("x1",function(d){return d.source.x;}).attr("y1",function(d){return d.source.y;})
            .attr("x2",function(d){return d.target.x;}).attr("y2",function(d){return d.target.y;});
        linkLabel.attr("x",function(d){return (d.source.x+d.target.x)/2;}).attr("y",function(d){return (d.source.y+d.target.y)/2-6;});
        nodeG.attr("transform",function(d){return "translate("+d.x+","+d.y+")";});
    });
}

function closeDetail(){
    document.getElementById("detail-panel").classList.remove("open");
}
function showDetail(d){
    var c=COLORS[d.domain]||"#94a3b8";
    var p=document.getElementById("detail-content");
    var cols=COLS[d.id]||[];
    var h="<div class=\"dp-title\">"+d.id+"</div><div class=\"dp-meta\">"+d.bo+" <span style=\"background:"+c+"22;color:"+c+";border:1px solid "+c+"44\">"+d.domain+"</span></div><div class=\"dp-desc\">"+d.desc+"</div>";
    if(cols.length>0){
        h+="<div style=\"font-size:12px;color:#64748b;margin-bottom:8px\">comment on table "+d.id+" is '<span style=\"color:#e2e8f0\">"+d.desc+"</span>';</div>";
        h+="<div style=\"height:1px;background:rgba(255,255,255,0.06);margin:8px 0 12px 0\"></div>";
        cols.forEach(function(col){
            var val=(col.r&&col.r!==col.n&&col.r!==col.t)?col.r:col.t;
            h+="<div class=\"col-item\"><span class=\"kw\">comment on column </span><span class=\"tbl\">"+d.id+"."+col.n+"</span><span class=\"kw\"> is '</span><span class=\"val\">"+val+"</span><span class=\"kw\">';</span><span class=\"type\">-- "+col.tp+"</span></div>";
        });
    }else{
        h+="<div style=\"color:#475569;font-size:13px;padding:24px 0;text-align:center\">暂无字段信息</div>";
    }
    p.innerHTML = h;
    document.getElementById("detail-panel").classList.add("open");
}

window.addEventListener("resize",function(){
    if(!sim)return;
    var w=window.innerWidth,h=window.innerHeight;
    svg.attr("viewBox",[0,0,w,h]);
    sim.force("center",d3.forceCenter(w/2,h/2)).alpha(0.2).restart();
});
</script>
</body>
</html>'''

    # Write html
    html_path = os.path.join(outdir, "graph.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print("HTML:", html_path)

    # Write data.js
    data_path = os.path.join(outdir, "data.js")
    with open(data_path, "w", encoding="utf-8") as f:
        f.write("var NODES=" + nodes_json + ";\n")
        f.write("var EDGES=" + edges_json + ";\n")
        f.write("var COLS=" + cols_json_all + ";\n")
    print("Data:", data_path, f"({os.path.getsize(data_path)} bytes)")
    print("\nDone! Open graph.html directly in your browser.")

if __name__ == "__main__":
    main()
