
import openpyxl, os, sys, json

Q = chr(34)  # double quote
S = chr(39)  # single quote

def main():
    path = sys.argv[1]
    if os.path.isdir(path):
        fs = [f for f in os.listdir(path) if f.endswith('.xlsx') and not f.startswith('~$')]
        path = os.path.join(path, fs[0])
    outdir = os.path.dirname(os.path.abspath(path))
    
    wb = openpyxl.load_workbook(path)
    ws = wb['表目录']
    
    # Parse catalog
    rows = []; cn = cb = None
    for i in range(7, ws.max_row + 1):
        c4=ws.cell(i,4).value;c5=ws.cell(i,5).value;c6=ws.cell(i,6).value;c7=ws.cell(i,7).value;c8=ws.cell(i,8).value
        if c4:
            try: cn=int(str(c4).strip())
            except: pass
        if c5: cb=str(c5).strip()
        if c6 and c7: rows.append({"bo":cb,"desc":str(c6).strip(),"tbl":str(c7).strip(),"rel":str(c8).strip() if c8 else ""})
    
    # Build nodes/edges
    nids=set();nodes=[];edges=[]
    for r in rows:
        if r['tbl']in nids: continue
        nids.add(r['tbl'])
        nodes.append({"id":r['tbl'],"label":r['tbl'].replace('t_invp_','').replace('tk_',''),"bo":r['bo'],"desc":r['desc'],"domain":"U0001f535","type":"main"if not r['rel']else"child"})
    for r in rows:
        if r['rel']and r['rel']!=r['tbl']and r['tbl']in nids and r['rel']in nids:
            edges.append({"source":r['rel'],"target":r['tbl'],"label":r['desc']})
    
    # Get columns
    cols={}
    for sn in wb.sheetnames:
        if sn=='表目录': continue
        s=wb[sn];ct=None;c=[];h=False
        for i in range(1,s.max_row+1):
            c2=s.cell(i,2).value;c5=s.cell(i,5).value;c8=s.cell(i,8).value
            if c2 and str(c2).strip()=='数据表名：' and c5:
                if ct and c:
                    if ct not in cols or len(c)>len(cols.get(ct,[])): cols[ct]=c
                ct=str(c5).strip();c=[];h=False
            elif c2 and str(c2).strip()=='编号':
                c3=s.cell(i,3).value;c4=s.cell(i,4).value
                if c3 and c4 and str(c3).strip()=='列名' and '列标题'in str(c4): h=True
            elif h and c2:
                try:
                    float(str(c2).strip())
                    cn2=str(s.cell(i,3).value).strip() if s.cell(i,3).value else ''
                    ct2=str(s.cell(i,4).value).strip() if s.cell(i,4).value else cn2
                    tp=str(s.cell(i,5).value).strip() if s.cell(i,5).value else ''
                    rr=str(s.cell(i,11).value).strip() if s.cell(i,11).value else ''
                    if cn2: c.append({"n":cn2,"t":ct2,"tp":tp,"r":rr})
                except:
                    if c8 and str(c8).strip()=='返回主目录': h=False
        if ct and c:
            if ct not in cols or len(c)>len(cols.get(ct,[])): cols[ct]=c
    
    # JSON data
    j_idx=json.dumps([{"id":n["id"],"label":n["label"],"bo":n["bo"],"domain":n["domain"],"desc":n["desc"]} for n in nodes],ensure_ascii=False)
    j_nm=json.dumps({n["id"]:n for n in nodes},ensure_ascii=False)
    j_eg=json.dumps(edges,ensure_ascii=False)
    j_cl=json.dumps(cols,ensure_ascii=False)
    j_cr=json.dumps({"U0001f535":"#3b82f6","U0001f7e2":"#10b981","U0001f7e0":"#f59e0b","U0001f534":"#ef4444","⚪":"#6b7280","U0001f7e3":"#8b5cf6","其他":"#94a3b8"},ensure_ascii=False)
    
    # Build HTML using chr() for quoting to avoid ALL escaping issues
    Q = chr(34)
    SQ = chr(39)
    
    parts = []
    parts.append('<!DOCTYPE html>\n<html lang=' + Q + 'zh-CN' + Q + '>\n<head>\n<meta charset=' + Q + 'UTF-8' + Q + '><meta name=' + Q + 'viewport' + Q + ' content=' + Q + 'width=device-width,initial-scale=1.0' + Q + '>\n')
    parts.append('<title>数据字典 · 关联模型</title>\n')
    parts.append('<script src=' + Q + 'https://d3js.org/d3.v7.min.js' + Q + '></script>\n')
    parts.append('<style>')
    parts.append('*{margin:0;padding:0;box-sizing:border-box}')
    parts.append('body{font-family:-apple-system,BlinkMacSystemFont,' + Q + 'Segoe UI' + Q + ',Roboto,sans-serif;background:#0f172a;color:#e2e8f0;overflow:hidden}')
    parts.append('#hd{position:fixed;top:0;left:0;right:0;z-index:10;padding:12px 24px;background:rgba(15,23,42,0.9);backdrop-filter:blur(12px);border-bottom:1px solid rgba(255,255,255,0.06);display:flex;align-items:center;gap:12px}')
    parts.append('#hd h1{font-size:18px;font-weight:600}#hd span{font-size:13px;color:#94a3b8}')
    parts.append('#sr{position:fixed;top:56px;left:50%;transform:translateX(-50%);z-index:15;width:420px;padding:10px 0}')
    parts.append('#sr input{width:100%;padding:9px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.08);background:rgba(30,41,59,0.95);color:#e2e8f0;font-size:13px;outline:none;transition:border-color 0.2s;box-sizing:border-box}')
    parts.append('#sr input:focus{border-color:rgba(255,255,255,0.25)}#sr input::placeholder{color:#475569}')
    parts.append('#sg{position:absolute;top:48px;left:0;right:0;background:#1e293b;border:1px solid rgba(255,255,255,0.08);border-radius:8px;max-height:320px;overflow-y:auto;display:none}')
    parts.append('.si{padding:8px 14px;cursor:pointer;font-size:13px;border-bottom:1px solid rgba(255,255,255,0.03);display:flex;align-items:center;gap:8px;transition:background 0.15s}')
    parts.append('.si:hover{background:rgba(255,255,255,0.06)}.si .sl{font-family:monospace;color:#e2e8f0}.si .sb{color:#64748b;font-size:11px}')
    parts.append('.si .sd{font-size:10px;padding:1px 6px;border-radius:3px;margin-left:auto}')
    parts.append('#lg{position:fixed;bottom:24px;left:24px;z-index:10;display:flex;gap:8px;flex-wrap:wrap;background:rgba(30,41,59,0.9);backdrop-filter:blur(12px);padding:10px 14px;border-radius:10px;border:1px solid rgba(255,255,255,0.06)}')
    parts.append('.li{display:flex;align-items:center;gap:6px;font-size:12px;color:#cbd5e1}.ld{width:10px;height:10px;border-radius:50%}')
    parts.append('#tp{z-index:20;background:rgba(30,41,59,0.96);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:14px 18px;font-size:13px;line-height:1.6;max-width:400px;pointer-events:none;opacity:0;transition:opacity 0.15s;position:fixed;z-index:20}')
    parts.append('#tp.show{opacity:1}#tp .tt{font-size:15px;font-weight:600;color:#f1f5f9;word-break:break-all}#tp .tb{color:#94a3b8;font-size:12px}')
    parts.append('svg{position:fixed;top:0;left:0;width:100%;height:100%}')
    parts.append('.el{stroke:rgba(255,255,255,0.12);stroke-width:1.5;fill:none;transition:stroke 0.2s}')
    parts.append('.ec{stroke:rgba(255,255,255,0.2);stroke-width:1;stroke-dasharray:4,3}')
    parts.append('.ell{font-size:9px;fill:#64748b;pointer-events:none;text-shadow:0 0 4px #0f172a}')
    parts.append('.nc{cursor:pointer;transition:r 0.2s,opacity 0.2s}.nc:hover{filter:brightness(1.3)}')
    parts.append('.nl{font-size:11px;fill:#e2e8f0;pointer-events:none;font-family:monospace;text-shadow:0 0 4px rgba(0,0,0,0.8)}')
    parts.append('.nd{fill:#475569}.mk{fill:rgba(255,255,255,0.2)}')
    parts.append('#st{position:fixed;top:96px;left:50%;transform:translateX(-50%);z-index:5;font-size:13px;color:#475569;pointer-events:none;text-align:center}')
    parts.append('#dp{position:fixed;top:56px;right:0;bottom:0;width:460px;z-index:25;background:rgba(15,23,42,0.97);backdrop-filter:blur(20px);border-left:1px solid rgba(255,255,255,0.08);transform:translateX(100%);transition:transform 0.25s;overflow-y:auto;padding:24px}')
    parts.append('#dp.o{transform:translateX(0)}#dc{position:sticky;top:0;float:right;cursor:pointer;font-size:24px;color:#64748b;padding:2px 8px;line-height:1}')
    parts.append('.dt{font-size:16px;font-weight:600;color:#f1f5f9;word-break:break-all;margin-bottom:4px;padding-right:32px}')
    parts.append('.dm{font-size:12px;color:#64748b;margin-bottom:16px;line-height:1.8}.dm span{display:inline-block;padding:1px 8px;border-radius:4px;font-size:11px;margin-left:6px}')
    parts.append('.dd{color:#94a3b8;font-size:13px;margin-bottom:16px;padding:8px 12px;background:rgba(255,255,255,0.03);border-radius:6px}')
    parts.append('.ci{font-size:12px;font-family:monospace;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.03);line-height:1.7}')
    parts.append('.ci .c1{color:#64748b}.ci .c2{color:#e2e8f0}.ci .c3{color:#a5d6ff}.ci .c4{color:#475569;margin-left:8px;font-size:11px;font-family:sans-serif}')
    parts.append('#ci{position:fixed;bottom:24px;right:24px;z-index:10;font-size:12px;color:#64748b;background:rgba(30,41,59,0.8);backdrop-filter:blur(8px);padding:8px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.04)}')
    parts.append('</style>\n</head>\n<body>')
    parts.append('<div id=' + Q + 'hd' + Q + '><h1>数据字典 · 关联模型</h1><span>' + str(len(nodes)) + '表 · 搜索发现</span></div>')
    parts.append('<div id=' + Q + 'sr' + Q + '><input type=' + Q + 'text' + Q + ' id=' + Q + 'si' + Q + ' placeholder=' + Q + '搜索表名 / 业务对象...' + Q + ' autocomplete=' + Q + 'off' + Q + '><div id=' + Q + 'sg' + Q + '></div></div>')
    parts.append('<div id=' + Q + 'st' + Q + '>输入关键字搜索</div>')
    parts.append('<div id=' + Q + 'lg' + Q + '></div><div id=' + Q + 'tp' + Q + '></div>')
    parts.append('<div id=' + Q + 'dp' + Q + '><div id=' + Q + 'dc' + Q + ' onclick=' + Q + 'cl()' + Q + '>&times;</div><div id=' + Q + 'dc2' + Q + '></div></div>')
    parts.append('<div id=' + Q + 'ci' + Q + '>99 表</div>')
    parts.append('<svg id=' + Q + 'g' + Q + '></svg>')
    parts.append('<script>')
    parts.append('var IDX=' + j_idx + ';var NM=' + j_nm + ';var EG=' + j_eg + ';var CL=' + j_cl + ';var CR=' + j_cr + ';')
    parts.append('var lg=document.getElementById(' + SQ + 'lg' + SQ + ');')
    parts.append('var ds=[...new Set(IDX.map(function(d){return d.domain;}))];')
    parts.append('lg.innerHTML=ds.map(function(d){return' + SQ + '<span class="li"><span class="ld" style="background:' + SQ + '+CR[d]+' + SQ + '"></span>' + SQ + '+d+'+SQ+'</span>'+SQ+';}).join('+SQ+''+SQ+');')
    parts.append('var si=document.getElementById(' + SQ + 'si' + SQ + ');var sg=document.getElementById(' + SQ + 'sg' + SQ + ');var st=document.getElementById(' + SQ + 'st' + SQ + ');')
    parts.append('var tp=document.getElementById(' + SQ + 'tp' + SQ + ');')
    parts.append('si.addEventListener(' + SQ + 'input' + SQ + ',function(){sg.style.display=si.value.trim()?' + SQ + 'block' + SQ + ':' + SQ + 'none' + SQ + ';})')
    parts.append('</script>')
    parts.append('</body>\n</html>')
    
    result = '\n'.join(parts)
    with open(os.path.join(outdir, 'graph.html'), 'w', encoding='utf-8') as f:
        f.write(result)
    print('OK:', os.path.getsize(os.path.join(outdir, 'graph.html')), 'bytes')

main()
