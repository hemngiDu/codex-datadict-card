#!/usr/bin/env python3
"""
Scalable graph generator: search-first + on-demand loading
Generates: index.json, chunks/, columns/, graph.html
"""
import openpyxl, os, sys, json, glob

# Config
DOMAIN_GROUPS = [
    ("配置与方案",["库存计划方案","安全库存统计方案","日均消耗统计方案","调度方案"]),
    ("因子与水位",["库存水位信息","库存水位因子","因子数据维护","因子取数方案","因子取数规则","补货日期"]),
    ("算法与计算",["算法方案配置","算法注册配置","库存水位更新方案","智能计算设置","因子分批计算参数","快速搭建模型","资源注册模型"]),
    ("执行与结果",["库存计划建议","供需匹配明细","安全库存记录","匹配映射配置","日均消耗记录","客户服务水平"]),
    ("日志审计",["消耗量计算日志","安全库存计算日志","库存计划运算日志"]),
    ("辅助测试",["组织测试"]),
]
CROSS_REFS = [
    ("t_invp_scheme","t_invp_model_register","需求/供应来源模型"),
    ("t_invp_scheme","t_invp_algoconfig","算法方案"),
    ("t_invp_scheme","t_invp_calrulecfg","水位更新方案"),
    ("t_invp_scheme","t_invp_invlevel","库存水位信息"),
    ("t_invp_scheme","t_invp_matchconfig","匹配映射配置"),
    ("t_invp_planadvice","t_invp_scheme","所属计划方案"),
    ("t_invp_invlevel","t_invp_levelfactor","水位因子"),
    ("t_invp_matchdetail","t_invp_scheme","计划方案"),
    ("t_invp_smartcalccfg","t_invp_invlevel","库存水位"),
    ("t_invp_smartcalccfg","t_invp_calrulecfg","因子计算规则"),
    ("t_invp_smartcalccfg","t_invp_queryschema","因子取数方案"),
    ("t_invp_invleveldata","t_invp_levelfactor","水位因子"),
    ("t_invp_ssrecord","t_invp_ssdayscheme","安全库存方案"),
    ("t_invp_dac_record","t_invp_dailycomsumption","日均消耗方案"),
    ("t_invp_safestock_callog","t_invp_ssdayscheme","安全库存方案"),
]
DOMAIN_COLORS = {"配置与方案":"#3b82f6","因子与水位":"#10b981","算法与计算":"#f59e0b","执行与结果":"#ef4444","日志审计":"#6b7280","辅助测试":"#8b5cf6","其他":"#94a3b8"}

def get_domain(bo):
    for d, objs in DOMAIN_GROUPS:
        if bo in objs: return d
    return "其他"

def get_domain_color(domain):
    return DOMAIN_COLORS.get(domain, "#94a3b8")

def parse_catalog(wb):
    ws = wb["表目录"]
    rows = []
    curr_no, curr_bo = None, None
    for r in range(7, ws.max_row + 1):
        c4 = ws.cell(r,4).value; c5 = ws.cell(r,5).value
        c6 = ws.cell(r,6).value; c7 = ws.cell(r,7).value; c8 = ws.cell(r,8).value
        if c4:
            try: curr_no = int(str(c4).strip())
            except: pass
        if c5: curr_bo = str(c5).strip()
        if c6 and c7:
            rows.append({"no":curr_no,"bo":curr_bo,"desc":str(c6).strip(),"tbl":str(c7).strip(),"rel":str(c8).strip() if c8 else ""})
    return rows

def get_column_details(wb):
    details = {}
    for sn in wb.sheetnames:
        if sn == "表目录": continue
        sws = wb[sn]
        cur_tbl = None; cols = []; in_hdr = False
        for r in range(1, sws.max_row + 1):
            c2 = sws.cell(r,2).value; c5 = sws.cell(r,5).value; c8 = sws.cell(r,8).value
            if c2 and str(c2).strip() == "数据表名：" and c5:
                if cur_tbl and cols:
                    if cur_tbl not in details or len(cols) > len(details.get(cur_tbl, [])):
                        details[cur_tbl] = cols
                cur_tbl = str(c5).strip(); cols = []; in_hdr = False
            elif c2 and str(c2).strip() == "编号":
                c3 = sws.cell(r,3).value; c4 = sws.cell(r,4).value
                if c3 and c4 and str(c3).strip() == "列名" and "列标题" in str(c4):
                    in_hdr = True
            elif in_hdr and c2:
                try:
                    float(str(c2).strip())
                    cn = str(sws.cell(r,3).value).strip() if sws.cell(r,3).value else ""
                    ct = str(sws.cell(r,4).value).strip() if sws.cell(r,4).value else cn
                    tp = str(sws.cell(r,5).value).strip() if sws.cell(r,5).value else ""
                    rr = str(sws.cell(r,11).value).strip() if sws.cell(r,11).value else ""
                    if cn: cols.append({"n":cn,"t":ct,"tp":tp,"r":rr})
                except:
                    if c8 and str(c8).strip() == "返回主目录": in_hdr = False
        if cur_tbl and cols:
            if cur_tbl not in details or len(cols) > len(details.get(cur_tbl, [])):
                details[cur_tbl] = cols
    return details

def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_scalable.py <Excel file or dir>")
        sys.exit(1)
    path = sys.argv[1]
    if os.path.isdir(path):
        files = [f for f in os.listdir(path) if f.endswith(".xlsx") and not f.startswith("~$")]
        if not files: print("No xlsx found"); sys.exit(1)
        path = os.path.join(path, files[0])
    base = os.path.splitext(os.path.basename(path))[0]
    outdir = os.path.dirname(os.path.abspath(path))
    
    print("Reading:", path)
    wb = openpyxl.load_workbook(path)
    rows = parse_catalog(wb)
    print("Catalog:", len(rows), "tables")
    
    # Build nodes and edges
    node_ids = set(); nodes = []; edges = []
    for row in rows:
        if row["tbl"] in node_ids: continue
        node_ids.add(row["tbl"])
        d = get_domain(row["bo"])
        nt = "main" if not row["rel"] else "child"
        nodes.append({"id":row["tbl"],"label":row["tbl"].replace("t_invp_","").replace("tk_",""),"bo":row["bo"],"desc":row["desc"],"domain":d,"type":nt})
    for row in rows:
        if row["rel"] and row["rel"] != row["tbl"] and row["tbl"] in node_ids and row["rel"] in node_ids:
            edges.append({"source":row["rel"],"target":row["tbl"],"label":row["desc"]})
    for src, tgt, lbl in CROSS_REFS:
        if src in node_ids and tgt in node_ids:
            edges.append({"source":src,"target":tgt,"label":lbl,"type":"cross"})
    
    # Group by BO
    bo_tables = {}
    for n in nodes:
        bo = n["bo"]
        if bo not in bo_tables: bo_tables[bo] = []
        bo_tables[bo].append(n["id"])
    
    # Column data
    col_details = get_column_details(wb)
    
    # Create output dirs
    out_graph = os.path.join(outdir, "graph.html")
    out_index = os.path.join(outdir, "index.json")
    out_chunks = os.path.join(outdir, "chunks")
    out_columns = os.path.join(outdir, "columns")
    os.makedirs(out_chunks, exist_ok=True)
    os.makedirs(out_columns, exist_ok=True)
    
    # Save column files (per table)
    for tbl, cols in col_details.items():
        cf = os.path.join(out_columns, tbl + ".json")
        with open(cf, "w", encoding="utf-8") as f:
            json.dump(cols[:50], f, ensure_ascii=False)
    print("Columns:", len(col_details), "files")
    
    # Save chunk files (per BO)
    bo_nodes_map = {}; bo_edges_map = {}
    for bo, tbl_list in bo_tables.items():
        bo_key = bo.replace(" ","")
        bn = [n for n in nodes if n["bo"] == bo]
        tbl_set = set(tbl_list)
        be = [e for e in edges if (e["source"] in tbl_set and e["target"] in tbl_set)]
        cf = os.path.join(out_chunks, bo_key + ".json")
        with open(cf, "w", encoding="utf-8") as f:
            json.dump({"nodes":bn,"edges":be}, f, ensure_ascii=False)
        bo_nodes_map[bo] = bn
        bo_edges_map[bo] = be
    
    # Build search index
    index = []
    for n in nodes:
        entry = {"id":n["id"],"label":n["label"],"bo":n["bo"],"domain":n["domain"],"desc":n["desc"]}
        nm = n["bo"].replace(" ","")
        entry["chunk"] = nm
        index.append(entry)
    
    with open(out_index, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False)
    print("Index:", len(index), "entries")
    
    # Cross-ref edges for HTML
    xref_edges = [e for e in edges if e.get("type") == "cross"]
    xref_json = json.dumps(xref_edges, ensure_ascii=False)
    colors_json = json.dumps(DOMAIN_COLORS, ensure_ascii=False)
    
    # Generate HTML
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>数据字典 · 关联模型</title>
<script src="https://d3js.org/d3.v7.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0f172a;color:#e2e8f0;overflow:hidden}}
#header{{position:fixed;top:0;left:0;right:0;z-index:10;padding:12px 24px;background:rgba(15,23,42,0.9);backdrop-filter:blur(12px);border-bottom:1px solid rgba(255,255,255,0.06);display:flex;align-items:center;gap:12px}}
#header h1{{font-size:18px;font-weight:600}}
#header span{{font-size:13px;color:#94a3b8}}
#search-area{{position:fixed;top:56px;left:50%;transform:translateX(-50%);z-index:15;width:420px;padding:10px 0}}
#search-area input{{width:100%;padding:9px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.08);background:rgba(30,41,59,0.95);backdrop-filter:blur(12px);color:#e2e8f0;font-size:13px;outline:none;transition:border-color 0.2s;box-sizing:border-box}}
#search-area input:focus{{border-color:rgba(255,255,255,0.25)}}
#search-area input::placeholder{{color:#475569}}
#suggestions{{position:absolute;top:48px;left:0;right:0;background:rgba(30,41,59,0.98);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.08);border-radius:8px;max-height:320px;overflow-y:auto;display:none}}
.sg-item{{padding:8px 14px;cursor:pointer;font-size:13px;border-bottom:1px solid rgba(255,255,255,0.03);display:flex;align-items:center;gap:8px;transition:background 0.15s}}
.sg-item:hover{{background:rgba(255,255,255,0.06)}}
.sg-item .sg-label{{font-family:monospace;color:#e2e8f0}}
.sg-item .sg-bo{{color:#64748b;font-size:11px}}
.sg-item .sg-domain{{font-size:10px;padding:1px 6px;border-radius:3px;margin-left:auto}}
#legend{{position:fixed;bottom:24px;left:24px;z-index:10;display:flex;gap:8px;flex-wrap:wrap;background:rgba(30,41,59,0.9);backdrop-filter:blur(12px);padding:10px 14px;border-radius:10px;border:1px solid rgba(255,255,255,0.06)}}
.legend-item{{display:flex;align-items:center;gap:6px;font-size:12px;color:#cbd5e1}}
.legend-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}
#tooltip{{position:fixed;z-index:20;background:rgba(30,41,59,0.96);backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.1);border-radius:10px;padding:14px 18px;font-size:13px;line-height:1.6;max-width:400px;pointer-events:none;opacity:0;transition:opacity 0.15s;box-shadow:0 8px 32px rgba(0,0,0,0.4)}}
#tooltip.show{{opacity:1}}
#tooltip .tt-title{{font-size:15px;font-weight:600;color:#f1f5f9;margin-bottom:4px;word-break:break-all}}
#tooltip .tt-bo{{color:#94a3b8;font-size:12px;margin-bottom:6px}}
svg{{position:fixed;top:0;left:0;width:100%;height:100%}}
.edge-line{{stroke:rgba(255,255,255,0.12);stroke-width:1.5;fill:none;transition:stroke 0.2s}}
.edge-cross{{stroke:rgba(255,255,255,0.18);stroke-width:1;stroke-dasharray:4,3}}
.edge-label{{font-size:9px;fill:#64748b;pointer-events:none;text-shadow:0 0 4px #0f172a}}
.node-circle{{cursor:pointer;transition:r 0.2s,opacity 0.2s}}
.node-circle:hover{{filter:brightness(1.3)}}
.node-label{{font-size:11px;fill:#e2e8f0;pointer-events:none;font-family:monospace;text-shadow:0 0 4px rgba(0,0,0,0.8)}}
.node-label.dim{{fill:#475569}}
.marker{{fill:rgba(255,255,255,0.2)}}
#status{{position:fixed;top:96px;left:50%;transform:translateX(-50%);z-index:5;font-size:13px;color:#475569;pointer-events:none}}
#detail-panel{{position:fixed;top:56px;right:0;bottom:0;width:480px;z-index:25;background:rgba(15,23,42,0.97);backdrop-filter:blur(20px);border-left:1px solid rgba(255,255,255,0.08);transform:translateX(100%);transition:transform 0.25s cubic-bezier(0.4,0,0.2,1);overflow-y:auto;padding:24px}}
#detail-panel.open{{transform:translateX(0)}}
#detail-close{{position:sticky;top:0;float:right;cursor:pointer;font-size:24px;color:#64748b;padding:2px 8px;border-radius:4px;line-height:1}}
#detail-close:hover{{color:#e2e8f0;background:rgba(255,255,255,0.06)}}
.dp-title{{font-size:16px;font-weight:600;color:#f1f5f9;word-break:break-all;margin-bottom:4px;padding-right:32px}}
.dp-meta{{font-size:12px;color:#64748b;margin-bottom:6px;line-height:1.8}}
.dp-meta span{{display:inline-block;padding:1px 8px;border-radius:4px;font-size:11px;margin-left:6px}}
.dp-desc{{color:#94a3b8;font-size:13px;margin-bottom:16px;padding:8px 12px;background:rgba(255,255,255,0.03);border-radius:6px}}
.dp-loading{{color:#475569;font-size:13px;padding:24px 0;text-align:center}}
.col-item{{font-size:12px;font-family:monospace;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.03);line-height:1.7}}
.col-item .kw{{color:#64748b}}
.col-item .tbl{{color:#e2e8f0}}
.col-item .val{{color:#a5d6ff}}
.col-item .type{{color:#475569;margin-left:8px;font-size:11px;font-family:sans-serif}}
#count-info{{position:fixed;bottom:24px;right:24px;z-index:10;font-size:12px;color:#64748b;background:rgba(30,41,59,0.8);backdrop-filter:blur(8px);padding:8px 14px;border-radius:8px;border:1px solid rgba(255,255,255,0.04)}}
</style>
</head>
<body>
<div id="header"><h1>数据字典 · 关联模型</h1><span>搜索发现 · 按需加载</span></div>
<div id="search-area">
  <input type="text" id="searchInput" placeholder="搜索表名 / 业务对象..." autocomplete="off">
  <div id="suggestions"></div>
</div>
<div id="status">输入关键字搜索，匹配结果将自动加载关联子图</div>
<div id="legend"></div>
<div id="tooltip"></div>
<div id="detail-panel"><div id="detail-close" onclick="closeDetail()">&times;</div><div id="detail-content"></div></div>
<div id="count-info">0 表</div>
<svg id="graph"></svg>
<script>
const domainColors = ''' + colors_json + r''';
const xrefEdges = ''' + xref_json + r''';
let loadedChunks = new Set();
let loadedNodeData = new Set();
let indexData = [];
let allNodes = [];
let allEdges = [];
let sim = null;
let nodeG = null, link = null, linkLabel = null, g = null, svg = null;
let currentVisibleNodes = new Set();

// Load index on start
fetch("index.json").then(r=>r.json()).then(data=>{
    indexData = data;
    document.getElementById("count-info").textContent = data.length + " 表";
    buildLegend();
    // Check URL for search param
    const params = new URLSearchParams(window.location.search);
    const q = params.get("q");
    if(q){ document.getElementById("searchInput").value = q; doSearch(q); }
});

const legendEl = document.getElementById("legend");
function buildLegend(){
    const domains = [...new Set(indexData.map(d=>d.domain))];
    legendEl.innerHTML = domains.map(d=>
        "<div class=\"legend-item\"><span class=\"legend-dot\" style=\"background:"+(domainColors[d]||"#94a3b8")+"\"></span>"+d+"</div>"
    ).join("");
}

let searchTimer = null;
document.getElementById("searchInput").addEventListener("input", function(){
    clearTimeout(searchTimer);
    const v = this.value.trim();
    if(!v){ document.getElementById("suggestions").style.display="none"; return; }
    searchTimer = setTimeout(()=>showSuggestions(v), 150);
});

function showSuggestions(q){
    const ql = q.toLowerCase();
    const matches = indexData.filter(d=>
        d.id.toLowerCase().includes(ql) || d.label.toLowerCase().includes(ql) ||
        d.bo.toLowerCase().includes(ql) || d.desc.toLowerCase().includes(ql)
    ).slice(0, 30);
    const el = document.getElementById("suggestions");
    if(matches.length===0){ el.style.display="none"; return; }
    el.innerHTML = matches.map(d=>{
        const c = domainColors[d.domain]||"#94a3b8";
        return "<div class=\"sg-item\" onclick=\"selectItem('"+d.id+"','"+d.chunk+"')\">"+
            "<span class=\"sg-label\">"+d.label+"</span>"+
            "<span class=\"sg-bo\">"+d.bo+"</span>"+
            "<span class=\"sg-domain\" style=\"background:"+c+"22;color:"+c+"\">"+d.domain+"</span></div>";
    }).join("");
    el.style.display = "block";
}

function selectItem(id, chunk){
    document.getElementById("suggestions").style.display="none";
    document.getElementById("searchInput").value = id;
    doSearch(id);
}

function doSearch(q){
    const ql = q.toLowerCase();
    const matchIds = new Set();
    indexData.forEach(d=>{
        if(d.id.toLowerCase().includes(ql) || d.label.toLowerCase().includes(ql) ||
           d.bo.toLowerCase().includes(ql) || d.desc.toLowerCase().includes(ql)){
            matchIds.add(d.id);
        }
    });
    if(matchIds.size===0){ document.getElementById("status").textContent="无匹配结果"; return; }
    
    // Find which chunks to load
    const neededChunks = new Set();
    matchIds.forEach(id=>{
        const entry = indexData.find(d=>d.id===id);
        if(entry) neededChunks.add(entry.chunk);
    });
    
    document.getElementById("status").textContent = "加载 " + neededChunks.size + " 个业务对象...";
    loadChunks(Array.from(neededChunks), matchIds);
}

function loadChunks(chunks, matchIds){
    let pending = chunks.length;
    chunks.forEach(chunk=>{
        if(loadedChunks.has(chunk)){ pending--; if(pending===0) renderGraph(matchIds); return; }
        fetch("chunks/"+chunk+".json").then(r=>r.json()).then(data=>{
            loadedChunks.add(chunk);
            data.nodes.forEach(n=>{ if(!allNodes.find(x=>x.id===n.id)) allNodes.push(n); });
            data.edges.forEach(e=>{ if(!allEdges.find(x=>x.source===e.source&&x.target===e.target)) allEdges.push(e); });
            pending--;
            if(pending===0) renderGraph(matchIds);
        }).catch(()=>{ pending--; if(pending===0) renderGraph(matchIds); });
    });
}

function renderGraph(matchIds){
    document.getElementById("status").textContent = "";
    currentVisibleNodes = new Set(matchIds);
    
    // Find 1-hop neighbors
    const visibleIds = new Set(matchIds);
    allEdges.forEach(e=>{
        const src = typeof e.source==="object"?e.source.id:e.source;
        const tgt = typeof e.target==="object"?e.target.id:e.target;
        if(matchIds.has(src)) visibleIds.add(tgt);
        if(matchIds.has(tgt)) visibleIds.add(src);
    });
    
    // Filter nodes & edges
    const visNodes = allNodes.filter(n=>visibleIds.has(n.id));
    const visEdges = allEdges.filter(e=>{
        const src = typeof e.source==="object"?e.source.id:e.source;
        const tgt = typeof e.target==="object"?e.target.id:e.target;
        return visibleIds.has(src) && visibleIds.has(tgt);
    });
    
    // Add cross-ref edges that connect visible nodes
    xrefEdges.forEach(e=>{
        if(visibleIds.has(e.source) && visibleIds.has(e.target)){
            if(!visEdges.find(x=>x.source===e.source&&x.target===e.target)){
                visEdges.push(e);
            }
        }
    });
    
    // Clean up old graph
    d3.select("svg").selectAll("*:not(defs)").remove();
    
    const width = window.innerWidth, height = window.innerHeight;
    svg = d3.select("#graph").attr("viewBox",[0,0,width,height]);
    g = svg.append("g");
    
    svg.call(d3.zoom().scaleExtent([0.15,4]).on("zoom",e=>g.attr("transform",e.transform))).on("dblclick.zoom",null);
    
    // Assign IDs to edges for marker references
    visEdges.forEach((e,i)=>{ e._idx = i; });
    
    const defs = svg.append("defs");
    visEdges.forEach((e,i)=>{
        defs.append("marker").attr("id","a"+i).attr("viewBox","0 -5 10 10")
            .attr("refX",20).attr("refY",0).attr("markerWidth",6).attr("markerHeight",6).attr("orient","auto")
            .append("path").attr("d","M0,-5L10,0L0,5").attr("class","marker");
    });
    
    sim = d3.forceSimulation(visNodes)
        .force("link",d3.forceLink(visEdges).id(d=>d.id).distance(d=>d.type==="cross"?200:130))
        .force("charge",d3.forceManyBody().strength(-350))
        .force("center",d3.forceCenter(width/2,height/2))
        .force("collision",d3.forceCollide(35));
    
    link = g.append("g").selectAll("line").data(visEdges).join("line")
        .attr("class",d=>d.type==="cross"?"edge-line edge-cross":"edge-line")
        .attr("marker-end",d=>"url(#a"+d._idx+")");
    
    linkLabel = g.append("g").selectAll("text").data(visEdges).join("text")
        .attr("class","edge-label").text(d=>d.label.substring(0,14));
    
    nodeG = g.append("g").selectAll("g").data(visNodes).join("g")
        .call(d3.drag().on("start",(e,d)=>{if(!e.active)sim.alphaTarget(0.3).restart();d.fx=d.x;d.fy=d.y;})
            .on("drag",(e,d)=>{d.fx=e.x;d.fy=e.y;})
            .on("end",(e,d)=>{if(!e.active)sim.alphaTarget(0);d.fx=null;d.fy=null;}));
    
    nodeG.append("circle").attr("class","node-circle")
        .attr("r",d=>d.type==="main"?11:7)
        .attr("fill",d=>domainColors[d.domain]||"#94a3b8")
        .attr("stroke","rgba(255,255,255,0.15)").attr("stroke-width",1.5);
    
    nodeG.append("text").attr("class","node-label")
        .attr("dx",d=>d.type==="main"?15:12).attr("dy",4)
        .text(d=>d.label);
    
    // Tooltip
    const tip = document.getElementById("tooltip");
    
    function showTooltip(e,d){
        const c=domainColors[d.domain]||"#94a3b8";
        tip.innerHTML="<div class=\"tt-title\">"+d.id+"</div><div class=\"tt-bo\">"+d.bo+"</div><div class=\"tt-desc\">"+d.desc+"</div><span class=\"tt-domain\" style=\"background:"+c+"22;color:"+c+";border:1px solid "+c+"44\">"+d.domain+"</span>";
        let x=e.pageX+16,y=e.pageY-40;
        if(x+400>window.innerWidth)x=e.pageX-416;
        if(y<60)y=60;
        tip.style.left=x+"px";tip.style.top=y+"px";tip.classList.add("show");
        const cn=new Set();
        visEdges.forEach(ed=>{
            const s=typeof ed.source==="object"?ed.source.id:ed.source;
            const t=typeof ed.target==="object"?ed.target.id:ed.target;
            if(s===d.id||t===d.id){cn.add(s);cn.add(t);}
        });
        nodeG.selectAll("circle").attr("opacity",n=>cn.has(n.id)||n.id===d.id?1:0.08);
        nodeG.selectAll("text").attr("class",n=>cn.has(n.id)||n.id===d.id?"node-label":"node-label dim");
        link.style("opacity",ed=>{
            const s=typeof ed.source==="object"?ed.source.id:ed.source;
            const t=typeof ed.target==="object"?ed.target.id:ed.target;
            return(s===d.id||t===d.id)?1:0.04;
        });
    }
    function hideTooltip(){
        tip.classList.remove("show");
        nodeG.selectAll("circle").attr("opacity",1);
        nodeG.selectAll("text").attr("class","node-label");
        link.style("opacity",1);
    }
    nodeG.on("mouseenter",showTooltip).on("mousemove",showTooltip).on("mouseleave",hideTooltip);
    
    // Click to show details
    nodeG.on("click", function(e,d){ showDetail(d); });
    
    sim.on("tick",()=>{
        link.attr("x1",d=>d.source.x).attr("y1",d=>d.source.y).attr("x2",d=>d.target.x).attr("y2",d=>d.target.y);
        linkLabel.attr("x",d=>(d.source.x+d.target.x)/2).attr("y",d=>(d.source.y+d.target.y)/2-6);
        nodeG.attr("transform",d=>"translate("+d.x+","+d.y+")");
    });
    
    document.getElementById("count-info").textContent = visNodes.length + " 表 · " + matchIds.size + " 匹配";
}

function closeDetail(){
    document.getElementById("detail-panel").classList.remove("open");
}

function showDetail(d){
    const c=domainColors[d.domain]||"#94a3b8";
    const panel = document.getElementById("detail-content");
    panel.innerHTML = "<div class=\"dp-title\">"+d.id+"</div><div class=\"dp-meta\">"+d.bo+" <span style=\"background:"+c+"22;color:"+c+";border:1px solid "+c+"44\">"+d.domain+"</span></div><div class=\"dp-desc\">"+d.desc+"</div><div class=\"dp-loading\">加载字段信息...</div>";
    document.getElementById("detail-panel").classList.add("open");
    
    // Fetch column data on demand
    fetch("columns/"+d.id+".json").then(r=>r.json()).then(cols=>{
        let h = "<div style=\"font-size:12px;color:#64748b;margin-bottom:8px\">comment on table "+d.id+" is '<span style=\"color:#e2e8f0\">"+d.desc+"</span>';</div>";
        h += "<div style=\"height:1px;background:rgba(255,255,255,0.06);margin:8px 0 12px 0\"></div>";
        if(cols.length>0){
            cols.forEach(col=>{
                const val = (col.r&&col.r!==col.n&&col.r!==col.t)?col.r:col.t;
                h += "<div class=\"col-item\"><span class=\"kw\">comment on column </span><span class=\"tbl\">"+d.id+"."+col.n+"</span><span class=\"kw\"> is '</span><span class=\"val\">"+val+"</span><span class=\"kw\">';</span><span class=\"type\">-- "+col.tp+"</span></div>";
            });
        }else{
            h += "<div style=\"color:#475569;font-size:13px;padding:24px 0;text-align:center\">暂无字段信息</div>";
        }
        panel.innerHTML = h;
    }).catch(()=>{
        panel.innerHTML = "<div style=\"color:#475569;font-size:13px;padding:24px 0;text-align:center\">加载失败</div>";
    });
}

window.addEventListener("resize",()=>{
    if(!sim)return;
    const w=window.innerWidth,h=window.innerHeight;
    svg.attr("viewBox",[0,0,w,h]);
    sim.force("center",d3.forceCenter(w/2,h/2)).alpha(0.2).restart();
});
</script>
</body>
</html>'''
    
    # Write HTML (replace curly brackets that conflict with f-string)
    out_html = html.replace("\n", "
")
    # Fix: use actual newlines
    with open(out_graph, "w", encoding="utf-8") as f:
        f.write(html)
    
    print("HTML:", out_graph)
    print("Index:", out_index, f"({len(index)} entries)")
    print("Chunks:", len(bo_tables))
    print("Columns:", len(col_details))
    print("\nDone! Open graph.html and search.")

if __name__ == "__main__":
    main()
