import sys
fp = sys.argv[1] if len(sys.argv) > 1 else None
if not fp or not __import__('os').path.exists(fp):
    print("用法: python script.py <Excel文件路径>"); sys.exit(1)
import openpyxl, os
wb = openpyxl.load_workbook(fp)
output = []
base = os.path.splitext(os.path.basename(fp))[0]
outdir = os.path.dirname(os.path.abspath(fp))
def flush(tbl, desc, cols):
    """Return (table_comment_lines, column_comment_lines) for one table"""
    lines = []
    if tbl:
        d = desc.replace("'", "''") if desc else ""
        lines.append("comment on table {} is '{}';".format(tbl, d))
        for cn, ct, cr in cols:
            c = ct
            if c == cn and cr and cr.strip():
                c = cr.strip()
            c = c.replace("'", "''")
            lines.append("comment on column {}.{} is '{}';".format(tbl, cn, c))
        lines.append("")
    return lines

for sn in wb.sheetnames:
    if sn == "表目录":
        continue
    ws = wb[sn]
    tbl, desc, cols, in_cols = None, None, [], False
    
    for r in range(1, ws.max_row + 1):
        # col layout:
        # 1=empty, 2=编号/index, 3=列名, 4=列标题, 5=类型/表名值, 6=长度, 7=精度, 
        # 8=主键/返回主目录, 9=非空, 10=默认值, 11=备注
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value
        c8 = ws.cell(r, 8).value
        c11 = ws.cell(r, 11).value
        
        c2s = str(c2).strip() if c2 else ""
        c8s = str(c8).strip() if c8 else ""
        c5s = str(c5).strip() if c5 else ""
        
        if c2s == "表名：" and c5:
            # new table block - flush previous if any
            if cols:
                output.extend(flush(tbl, desc, cols))
            desc = c5s
            tbl = None
            cols = []
            in_cols = False
        elif c2s == "数据表名：" and c5:
            tbl = c5s
        elif c2s == "编号" and c3 and c4:
            c3s = str(c3).strip() if c3 else ""
            c4s = str(c4).strip() if c4 else ""
            if c3s == "列名" and "列标题" in c4s:
                # header row
                in_cols = True
                cols = []
        elif in_cols and c2:
            # try to see if c2 is a number (index column)
            try:
                float(str(c2).strip())
                # yes, it's a column data row
                cn = str(c3).strip() if c3 else ""
                ct = str(c4).strip() if c4 else cn
                cr = str(c11).strip() if c11 else ""
                if cn:
                    cols.append((cn, ct, cr))
            except ValueError:
                # not a number - might be 返回主目录 or something else
                if c8s == "返回主目录":
                    output.extend(flush(tbl, desc, cols))
                    tbl, desc, cols, in_cols = None, None, [], False
        elif c8s == "返回主目录":
            output.extend(flush(tbl, desc, cols))
            tbl, desc, cols, in_cols = None, None, [], False
    
    # end of sheet - flush remaining
    output.extend(flush(tbl, desc, cols))

result = "\n".join(output)
outpath = os.path.join(outdir, base + "_comments.sql")
with open(outpath, "w", encoding="utf-8") as f:
    f.write(result)
print("Total SQL lines: " + str(len(output)))
print("---First 20 lines---")
for line in output[:20]:
    print(line)
print("---")
print("Total character count: " + str(len(result)))
print(f"SQL: {outpath} ({len(output)} 行)")